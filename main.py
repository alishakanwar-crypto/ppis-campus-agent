"""
PPIS Campus Agent — Local Windows application that connects to Hikvision DVRs
on the school LAN and communicates with the cloud bot via WebSocket.

Features:
- Web-based local UI for DVR configuration, Excel upload, camera mapping
- Hikvision ISAPI integration for snapshot capture
- WebSocket client to cloud bot for receiving snapshot requests
- On-demand child photo capture and delivery
"""

from __future__ import annotations

import concurrent.futures
import contextvars
import faulthandler
import threading
faulthandler.enable()  # Print C-level crash tracebacks

import asyncio
import base64
import io
import json
import logging
import logging.handlers
import os
import re
import subprocess
import sys
import time
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

# Suppress Windows crash dialogs so the process exits silently on errors
# (allows run_forever.bat to auto-restart without a popup blocking it)
if sys.platform == "win32":
    try:
        import ctypes
        # SEM_FAILCRITICALERRORS=1 | SEM_NOGPFAULTERRORBOX=2 | SEM_NOOPENFILEERRORBOX=0x8000
        ctypes.windll.kernel32.SetErrorMode(0x8003)  # type: ignore[attr-defined]
    except Exception:
        pass

# --- Kill stale process on port 8897 BEFORE loading heavy DLLs ---
# Must happen before face_recognition/dlib/cv2 imports because
# subprocess.run(shell=True) crashes when those DLLs are loaded.
def _kill_port_holder_early(port: int = 8897) -> None:
    if sys.platform != "win32":
        return
    # CREATE_NO_WINDOW prevents cmd.exe from flashing a black window
    _no_win = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
    try:
        result = subprocess.run(
            f'netstat -ano | findstr :{port} | findstr LISTENING',
            capture_output=True, text=True, shell=True,
            creationflags=_no_win,
        )
        for line in result.stdout.strip().splitlines():
            parts = line.split()
            if parts:
                pid = parts[-1]
                if pid != str(os.getpid()):
                    subprocess.run(f"taskkill /F /PID {pid}",
                                   shell=True, capture_output=True,
                                   creationflags=_no_win)
    except Exception:
        pass

if __name__ == "__main__":
    from campus_instance import (
        DUPLICATE_INSTANCE_EXIT_CODE,
        acquire_single_instance,
    )

    if not acquire_single_instance():
        raise SystemExit(DUPLICATE_INSTANCE_EXIT_CODE)

    # Early diagnostic: write startup timestamp so we know the process launched
    try:
        _diag_path = Path(__file__).parent / "startup_diag.log"
        with open(_diag_path, "a", encoding="utf-8") as _df:
            _df.write(f"\n[{datetime.now().isoformat()}] main.py starting (PID={os.getpid()})...\n")
    except Exception:
        pass

import httpx
import websockets
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles

# --- dlib/numpy ABI compatibility check ---
# dlib compiled against numpy 1.x rejects numpy 2.x arrays with
# "Unsupported image type, must be 8bit gray or RGB image."
# Auto-fix: downgrade numpy to last 1.x release.
def _ensure_dlib_compat():
    try:
        import numpy as _np
        import dlib as _dlib
        _det = _dlib.get_frontal_face_detector()
        _test = _np.zeros((100, 100, 3), dtype=_np.uint8)
        _det(_test, 0)  # should not raise
    except ImportError:
        pass  # dlib not installed — nothing to check
    except Exception as e:
        if "Unsupported image type" in str(e):
            print(f"[AUTOFIX] dlib/numpy ABI mismatch: {e}")
            print("[AUTOFIX] Installing compatible numpy (1.26.4)...")
            import subprocess
            _nw = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            result = subprocess.run(
                [sys.executable, "-m", "pip", "install", "numpy==1.26.4"],
                capture_output=True, text=True,
                creationflags=_nw if sys.platform == "win32" else 0,
            )
            print(result.stdout[-500:] if result.stdout else "")
            if result.returncode == 0:
                print("[AUTOFIX] numpy fixed. Restarting...")
                sys.exit(42)  # run_forever.bat will auto-restart
            else:
                print(f"[AUTOFIX] pip failed: {result.stderr[-300:]}")

_ensure_dlib_compat()

from attendance_engine import engine as attendance_engine
import face_db
import recorder_auth
from mood_detector import MoodDetector
from teacher_sighting import TeacherSightingTracker

try:
    from PIL import Image
except ImportError:
    Image = None

_LOG_FORMAT = "%(asctime)s [%(levelname)s] %(name)s: %(message)s"
_LOG_FILE = Path(__file__).parent / "campus_agent.log"

logging.basicConfig(
    level=logging.INFO,
    format=_LOG_FORMAT,
    handlers=[
        logging.StreamHandler(),
        logging.handlers.RotatingFileHandler(
            _LOG_FILE, maxBytes=10 * 1024 * 1024, backupCount=3,
            encoding="utf-8"),
    ],
)
logger = logging.getLogger("ppis-agent")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
CLOUD_API_BASE = "https://ppis-whatsapp-bot.fly.dev"

# Mood detection and teacher sighting trackers
mood_detector = MoodDetector(cloud_url=CLOUD_API_BASE)
sighting_tracker = TeacherSightingTracker(cloud_url=CLOUD_API_BASE)
CONFIG_FILE = Path(__file__).parent / "config.json"
SNAPSHOT_DIR = Path(__file__).parent / "snapshots"
SNAPSHOT_DIR.mkdir(exist_ok=True)


async def fetch_config_from_cloud() -> dict | None:
    """Fetch full config from the cloud-hosted SQLite database.
    Returns None if cloud is unreachable."""
    url = f"{CLOUD_API_BASE}/api/agent-config/full"
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(url)
            if resp.status_code == 200:
                data = resp.json()
                logger.info(
                    f"Fetched config from cloud: "
                    f"{len(data.get('dvrs', []))} DVRs, "
                    f"{len(data.get('camera_mapping', {}))} camera mappings"
                )
                return data
            logger.warning(f"Cloud config API returned {resp.status_code}")
    except Exception as e:
        logger.warning(f"Could not fetch cloud config: {e}")
    return None


def load_config_local() -> dict:
    """Load config from local config.json (fallback)."""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
        except (json.JSONDecodeError, ValueError) as e:
            logging.getLogger("ppis-agent").warning(
                f"config.json is corrupted ({e}), using defaults"
            )
    return {
        "cloud_bot_url": "wss://ppis-whatsapp-bot.fly.dev/ws/agent",
        "agent_secret": os.environ.get("AGENT_SECRET", ""),
        "dvrs": [],
        "camera_mapping": {},
        "snapshot_dir": "snapshots",
        "local_port": 8897,
    }


def save_config(cfg: dict):
    """Save config to local config.json (cache for offline use)."""
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


def cleanup_junk_face_entries():
    """Remove known junk/duplicate face entries from local DB on startup.

    These entries were registered with incorrect names and cause
    false-positive matches and duplicate notifications.
    """
    import database as db_mod

    junk_person_ids = [
        "TEACHER_ALISHA",              # Incomplete name — duplicate of TEACHER_ALISHA_AHUJA
        "TEACHER_RECOGNITION",         # Junk — someone registered with just "recognition"
        "TEACHER_PRITY_SHARMA_TREACHER",  # Typo — "Treacher" instead of "Teacher"
        "TEACHER_HARDIK_RAWAT_GRADE_4A",  # Has "Grade 4A" in name — registration error
    ]

    total_deleted = 0
    for pid in junk_person_ids:
        deleted = db_mod.delete_person_faces(pid)
        if deleted:
            total_deleted += deleted
            logger.info(f"Cleanup: removed {deleted} junk entry(s) for {pid}")

    if total_deleted:
        logger.info(f"Cleanup: removed {total_deleted} total junk face entry(s)")
    return total_deleted


async def sync_faces_from_cloud() -> int:
    """Download registered face images from cloud and register locally.

    Uses incremental sync: first fetches lightweight manifest (no images),
    compares with local DB, then downloads only missing faces one-by-one.
    This prevents OOM on the school PC by avoiding bulk image downloads.

    Returns the number of faces synced.
    """
    agent_secret = os.environ.get("AGENT_SECRET", "")
    headers = {"X-Agent-Secret": agent_secret} if agent_secret else {}
    try:
        import gc
        import database as db_mod

        async with httpx.AsyncClient(timeout=30) as client:
            # Step 1: Get manifest (metadata only — no images, ~1KB)
            manifest_url = f"{CLOUD_API_BASE}/api/face/manifest"
            resp = await client.get(manifest_url, headers=headers)
            if resp.status_code == 404:
                # Fallback to old /images endpoint if manifest not available
                return await _sync_faces_legacy(client, headers)
            if resp.status_code != 200:
                logger.warning(f"Cloud face sync: manifest returned {resp.status_code}")
                return 0
            manifest = resp.json()
            if not manifest:
                logger.info("Cloud face sync: no faces registered in cloud")
                return 0

            # Step 2: Find which faces we're missing locally
            existing = db_mod.get_all_face_encodings()
            existing_keys = {
                (r["person_id"], r.get("angle", ""))
                for r in existing
            }
            missing = [
                f for f in manifest
                if (f["person_id"], f["angle"]) not in existing_keys
            ]

            # Step 2b: Update phone numbers for existing faces
            # Build lookup of local person_id -> phone
            local_phones = {}
            for r in existing:
                pid = r["person_id"]
                ph = r.get("phone", "") or ""
                if pid not in local_phones or len(ph) > len(local_phones[pid]):
                    local_phones[pid] = ph
            # Check cloud manifest for updated phones
            phones_updated = 0
            for face_meta in manifest:
                pid = face_meta["person_id"]
                cloud_phone = face_meta.get("phone", "") or ""
                local_phone = local_phones.get(pid, "")
                if pid in local_phones and cloud_phone and cloud_phone != local_phone:
                    # Cloud has a newer/different phone — update locally
                    db_mod.update_face_phone(pid, cloud_phone)
                    local_phones[pid] = cloud_phone
                    phones_updated += 1
            if phones_updated:
                logger.info(f"Cloud face sync: updated phone numbers for {phones_updated} person(s)")

            if not missing:
                logger.info(f"Cloud face sync: all {len(manifest)} faces already local")
                return phones_updated

            logger.info(f"Cloud face sync: {len(missing)} new face(s) to download")

            # Step 3: Download missing faces ONE AT A TIME (low memory)
            synced = 0
            for face_meta in missing:
                face_id = face_meta["id"]
                person_id = face_meta["person_id"]
                try:
                    img_resp = await client.get(
                        f"{CLOUD_API_BASE}/api/face/image/{face_id}",
                        headers=headers,
                    )
                    if img_resp.status_code != 200:
                        logger.warning(f"Cloud face sync: image {face_id} returned {img_resp.status_code}")
                        continue
                    image_bytes = img_resp.content
                    # Encoding a face takes seconds of pure CPU: on the event
                    # loop it holds up the campus link, so a parent asking for
                    # a photo during a sync waits for the whole sync.
                    result = await asyncio.to_thread(
                        face_db.register_face,
                        person_id=person_id,
                        name=face_meta["name"],
                        role=face_meta["role"],
                        phone=face_meta["phone"],
                        angle=face_meta["angle"],
                        image_bytes=image_bytes,
                    )
                    del image_bytes
                    gc.collect()
                    if result.get("success"):
                        synced += 1
                        logger.info(f"Cloud face sync: registered {face_meta['name']} ({person_id})")
                    else:
                        logger.warning(f"Cloud face sync: failed {person_id}: {result.get('error')}")
                except (MemoryError, OSError) as e:
                    logger.error(f"Cloud face sync: MEMORY ERROR for {person_id}, skipping: {e}")
                    import gc as _gc; _gc.collect()
                    continue
                except Exception as e:
                    logger.warning(f"Cloud face sync: error downloading {person_id}: {e}")

            logger.info(f"Cloud face sync complete: {synced} new face(s) synced")
            return synced
    except Exception as e:
        logger.warning(f"Cloud face sync failed: {e}")
        return 0


async def _reload_faces_off_loop() -> None:
    """Rebuild the face caches without freezing the campus link.

    Reading every encoding out of the database takes tens of seconds, and on
    the event loop a parent's photo request is not even read from the socket
    until it ends — which is why the first request after a start timed out.
    """
    await asyncio.to_thread(attendance_engine.reload_faces)


async def _sync_faces_legacy(client: httpx.AsyncClient, headers: dict) -> int:
    """Fallback: download all faces via /api/face/images (old endpoint)."""
    import gc
    import database as db_mod

    url = f"{CLOUD_API_BASE}/api/face/images"
    resp = await client.get(url, headers=headers)
    if resp.status_code != 200:
        return 0
    faces = resp.json()
    if not faces:
        return 0
    existing = db_mod.get_all_face_encodings()
    existing_keys = {(r["person_id"], r.get("angle", "")) for r in existing}
    synced = 0
    for face_data in faces:
        person_id = face_data["person_id"]
        angle = face_data["angle"]
        if (person_id, angle) in existing_keys:
            continue
        image_bytes = base64.b64decode(face_data["image_base64"])
        result = await asyncio.to_thread(
            face_db.register_face,
            person_id=person_id,
            name=face_data["name"],
            role=face_data["role"],
            phone=face_data["phone"],
            angle=angle,
            image_bytes=image_bytes,
        )
        del image_bytes
        gc.collect()
        if result.get("success"):
            synced += 1
            existing_keys.add((person_id, angle))
    logger.info(f"Cloud face sync (legacy): {synced} new face(s)")
    return synced


async def load_config() -> dict:
    """Load config: try cloud first, fall back to local config.json."""
    cloud_cfg = await fetch_config_from_cloud()
    if cloud_cfg and cloud_cfg.get("dvrs"):
        # Merge cloud data into a usable config dict
        cfg = {
            "cloud_bot_url": cloud_cfg.get("cloud_bot_url", "wss://ppis-whatsapp-bot.fly.dev/ws/agent"),
            "agent_secret": cloud_cfg.get("agent_secret", os.environ.get("AGENT_SECRET", "")),
            "dvrs": cloud_cfg.get("dvrs", []),
            "camera_mapping": cloud_cfg.get("camera_mapping", {}),
            "local_port": int(cloud_cfg.get("settings", {}).get("local_port", 8897)),
        }
        # Cache locally for offline fallback
        save_config(cfg)
        logger.info("Config loaded from cloud DB (cached locally)")
        return cfg
    # Fallback to local
    logger.info("Using local config.json (cloud unavailable or empty)")
    return load_config_local()


# Config will be loaded async in lifespan; use local as placeholder
config = load_config_local()

# ---------------------------------------------------------------------------
# Hikvision ISAPI — Snapshot capture
# ---------------------------------------------------------------------------

def _jpeg_size(data: bytes) -> tuple[int, int]:
    """Pixel size of a JPEG, or (0, 0) when it cannot be read."""
    if Image is None:
        return 0, 0
    try:
        with Image.open(io.BytesIO(data)) as img:
            return img.width, img.height
    except Exception:
        return 0, 0


def _jpeg_pixels(data: bytes) -> int:
    """Pixel count of a JPEG, or 0 when it cannot be read."""
    width, height = _jpeg_size(data)
    return width * height


def compress_jpeg(data: bytes, max_bytes: int = 200_000, quality_start: int = 70) -> bytes:
    """Compress a JPEG image to fit within max_bytes.
    
    Uses Pillow if available, otherwise returns original data.
    """
    if Image is None or len(data) <= max_bytes:
        return data
    try:
        img = Image.open(io.BytesIO(data))
        # Resize if very large (>1920px on any side)
        max_dim = 1920
        if img.width > max_dim or img.height > max_dim:
            ratio = min(max_dim / img.width, max_dim / img.height)
            new_size = (int(img.width * ratio), int(img.height * ratio))
            img = img.resize(new_size, Image.LANCZOS)
        # Try decreasing quality until it fits
        quality = quality_start
        while quality >= 20:
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            result = buf.getvalue()
            if len(result) <= max_bytes:
                logger.info(f"Compressed image: {len(data)} -> {len(result)} bytes (q={quality})")
                return result
            quality -= 10
        # Return best effort
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=20, optimize=True)
        return buf.getvalue()
    except Exception as e:
        logger.warning(f"Image compression failed: {e}, using original")
        return data


_last_capture_error: str = ""
_digest_auth_cache: dict[tuple[str, str, str], httpx.DigestAuth] = {}


class _DvrCaptureLimiter:
    def __init__(
        self,
        limit: int,
        background_wait: float,
        background_limit: int | None = None,
    ):
        self._semaphore = asyncio.Semaphore(limit)
        self._background_limit = min(
            max(1, background_limit if background_limit is not None else limit - 1),
            max(1, limit - 1),
        )
        self._background_semaphore = asyncio.Semaphore(self._background_limit)
        self._condition = asyncio.Condition()
        self._live_waiting = 0
        self._live_active = 0
        self._background_wait = background_wait

    async def __aenter__(self):
        await self.acquire(False)
        return self

    async def __aexit__(self, *_args):
        await self.release(False)

    async def acquire(self, live: bool) -> None:
        if live:
            async with self._condition:
                self._live_waiting += 1
            acquired = False
            try:
                await self._semaphore.acquire()
                acquired = True
            finally:
                async with self._condition:
                    self._live_waiting -= 1
                    if acquired:
                        self._live_active += 1
                    self._condition.notify_all()
            if not acquired:
                raise asyncio.CancelledError
            return

        async with self._condition:
            try:
                await asyncio.wait_for(
                    self._condition.wait_for(
                        lambda: self._live_waiting == 0 and self._live_active == 0
                    ),
                    timeout=self._background_wait,
                )
            except asyncio.TimeoutError:
                pass
        await self._background_semaphore.acquire()
        try:
            await self._semaphore.acquire()
        except BaseException:
            self._background_semaphore.release()
            raise

    async def release(self, live: bool) -> None:
        self._semaphore.release()
        if live:
            async with self._condition:
                self._live_active -= 1
                self._condition.notify_all()
        else:
            self._background_semaphore.release()


_DVR_CAPTURE_LIMIT = max(
    1, int(os.environ.get("DVR_CAPTURE_CONCURRENCY", "6"))
)
_DVR_BACKGROUND_WAIT_SECONDS = max(
    0.1, float(os.environ.get("DVR_BACKGROUND_WAIT_SECONDS", "5"))
)
_DVR_BACKGROUND_LIMIT = min(
    max(1, int(os.environ.get("DVR_BACKGROUND_CONCURRENCY", "4"))),
    max(1, _DVR_CAPTURE_LIMIT - 1),
)
_dvr_capture_limiters: dict[str, _DvrCaptureLimiter] = {}
_dvr_capture_limiters_lock = asyncio.Lock()


async def _dvr_limiter(ip: str) -> _DvrCaptureLimiter:
    limiter = _dvr_capture_limiters.get(ip)
    if limiter is None:
        async with _dvr_capture_limiters_lock:
            limiter = _dvr_capture_limiters.get(ip)
            if limiter is None:
                limiter = _DvrCaptureLimiter(
                    _DVR_CAPTURE_LIMIT,
                    _DVR_BACKGROUND_WAIT_SECONDS,
                    _DVR_BACKGROUND_LIMIT,
                )
                _dvr_capture_limiters[ip] = limiter
    return limiter


async def _acquire_dvr_capture(ip: str, background: bool) -> _DvrCaptureLimiter:
    limiter = await _dvr_limiter(ip)
    await limiter.acquire(not background)
    return limiter


def _digest_auth(ip: str, user: str, password: str) -> httpx.DigestAuth:
    key = (ip, user, password)
    auth = _digest_auth_cache.get(key)
    if auth is None:
        auth = httpx.DigestAuth(user, password)
        _digest_auth_cache[key] = auth
    return auth


# ---------------------------------------------------------------------------
# RTSP Snapshot Fallback (for DVRs where ISAPI auth is broken)
# ---------------------------------------------------------------------------
_RTSP_FALLBACK_IPS: set[str] = {"192.168.0.13"}  # DVR 4 — ISAPI 401 but RTSP works
_RTSP_COOLDOWN_SECONDS = max(
    1.0, float(os.environ.get("RTSP_FAILURE_COOLDOWN_SECONDS", "120"))
)
_rtsp_cooldowns: dict[str, float] = {}
_rtsp_timeout_warning_logged = False
_rtsp_timeout_warning_lock = threading.Lock()
_live_dvr_clients: dict[str, httpx.AsyncClient] = {}
_live_capture_preferences: dict[tuple[str, int], tuple[str, int]] = {}
_live_capture_preference_age: dict[tuple[str, int], float] = {}
# Largest picture a channel has ever handed us, so a camera that simply cannot
# do 1080p is not re-probed on every parent request.
_live_capture_best_pixels: dict[tuple[str, int], int] = {}
_live_capture_size_logged: set[tuple[str, int]] = set()
# Hikvision returns a small default JPEG (704x480) unless the wanted size is
# asked for, which parents see as a blurred classroom photo.
_LIVE_SNAPSHOT_WIDTH = max(0, int(os.environ.get("SNAPSHOT_WIDTH", "1920")))
_LIVE_SNAPSHOT_HEIGHT = max(0, int(os.environ.get("SNAPSHOT_HEIGHT", "1080")))
# How long a fallback (smaller/sub-stream) capture stays preferred before the
# full-size main stream is tried again.
_LIVE_CAPTURE_FALLBACK_TTL_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_FALLBACK_TTL_SECONDS", "900"))
)
# How many differently-shaped pictures to collect from a channel we have never
# measured before settling on its sharpest one.
_LIVE_CAPTURE_PROBE_PICTURES = max(
    1, int(os.environ.get("SNAPSHOT_PROBE_PICTURES", "3"))
)
# Time the extra sharpness probes may take. A picture in hand always beats a
# sharper one the parent never receives.
_LIVE_CAPTURE_PROBE_BUDGET_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_PROBE_BUDGET_SECONDS", "2"))
)
# Longest a single door (one URL with one auth scheme) may take before it is
# abandoned for the next one. A recorder that never answers the full-size
# request otherwise eats the whole attempt, and every retry knocks on the same
# silent door again.
_LIVE_CAPTURE_DOOR_TIMEOUT_SECONDS = max(
    0.5, float(os.environ.get("SNAPSHOT_DOOR_TIMEOUT_SECONDS", "5"))
)
# How long a door that hung is tried last for.
_LIVE_CAPTURE_SLOW_DOOR_TTL_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_SLOW_DOOR_TTL_SECONDS", "300"))
)
_live_capture_slow_doors: dict[tuple[str, int], dict[int, float]] = {}
# A channel whose every snapshot door stayed silent: knocking on them again
# costs a parent ~20s before the video fallback even starts, so that channel
# goes straight to video until the doors are worth a fresh try.
_LIVE_CAPTURE_SILENT_CHANNEL_TTL_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_SILENT_CHANNEL_TTL_SECONDS", "300"))
)
_live_capture_silent_channels: dict[tuple[str, int], float] = {}
# How recently a recorder must have served a picture for a silent channel to
# count as "the recorder is busy" rather than "these doors are dead".
_LIVE_CAPTURE_BUSY_RECORDER_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_BUSY_RECORDER_SECONDS", "120"))
)
# How many load-suspected silences a channel is forgiven before it is treated
# as dead anyway, so a genuinely broken door cannot cost every parent its wait.
_LIVE_CAPTURE_BUSY_FORGIVENESS = max(
    1, int(os.environ.get("SNAPSHOT_BUSY_FORGIVENESS", "3"))
)
_live_capture_busy_silences: dict[tuple[str, int], int] = {}
# Parents asking for one classroom at the same moment share the capture that is
# already running, instead of each queueing more work on the recorder.
_LIVE_CAPTURE_SHARING = (
    os.environ.get("SNAPSHOT_SHARE_CAPTURES", "1").strip().lower()
    not in {"0", "false", "no"}
)
_live_capture_in_flight: dict[tuple[str, int], asyncio.Task] = {}
# A parent's photo is worth the bytes: WhatsApp accepts images up to 5 MB, so
# only squeeze quality when the picture is far bigger than that.
_LIVE_SNAPSHOT_MAX_BYTES = max(
    100_000, int(os.environ.get("SNAPSHOT_MAX_BYTES", "1500000"))
)
_LIVE_SNAPSHOT_JPEG_QUALITY = min(
    95, max(40, int(os.environ.get("SNAPSHOT_JPEG_QUALITY", "92")))
)
_live_dvr_client_lock = asyncio.Lock()
_live_request_deadline: contextvars.ContextVar[float | None] = (
    contextvars.ContextVar("live_request_deadline", default=None)
)
_live_request_classroom: contextvars.ContextVar[str] = contextvars.ContextVar(
    "live_request_classroom", default=""
)
# Where a live capture writes its own timing, so the cloud can see which stage
# of a parent's request was slow without reading the campus PC's log.
_live_capture_report: contextvars.ContextVar[dict | None] = contextvars.ContextVar(
    "live_capture_report", default=None
)


_PROCESS_STARTED_AT = datetime.now(timezone(timedelta(hours=5, minutes=30)))


def _running_commit() -> str:
    """Short git commit this process was started from, '' when unknown.

    A restart that fails to kill the old process leaves it serving stale code,
    which is indistinguishable from a bad fix unless the cloud can see this.
    """
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=str(Path(__file__).resolve().parent),
            capture_output=True,
            text=True,
            timeout=10,
        )
    except Exception:
        return ""
    return result.stdout.strip() if result.returncode == 0 else ""


def _process_started_at_ist() -> str:
    return _PROCESS_STARTED_AT.strftime("%d-%m-%Y %H:%M:%S IST")


_AUTO_UPDATE_CHECK_SECONDS = max(
    60.0, float(os.environ.get("AUTO_UPDATE_CHECK_SECONDS", "600"))
)
_AUTO_UPDATE_ENABLED = (
    os.environ.get("AUTO_UPDATE_ENABLED", "1").strip().lower()
    not in {"0", "false", "no"}
)
# Exiting only updates the agent when the wrapper is there to bring it back;
# started by hand, an exit would simply leave the campus without an agent.
_STARTED_BY_WRAPPER = os.environ.get("PPIS_WRAPPER", "") == "1"


def _git(*args: str) -> str:
    """Run a read-only git command in the agent's checkout, '' on failure."""
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=str(Path(__file__).resolve().parent),
            capture_output=True,
            text=True,
            timeout=60,
        )
    except Exception as exc:
        logger.debug("git %s failed: %s", " ".join(args), _exception_text(exc))
        return ""
    if result.returncode != 0:
        logger.debug(
            "git %s failed: %s", " ".join(args), result.stderr.strip()[:200]
        )
        return ""
    return result.stdout.strip()


def _pending_update_commit() -> str:
    """The commit on origin/main we are not running yet, '' when current."""
    _git("fetch", "origin", "main")
    remote = _git("rev-parse", "--short", "origin/main")
    local = _git("rev-parse", "--short", "HEAD")
    if not remote or not local or remote == local:
        return ""
    return remote


def _work_in_flight() -> str:
    """What the agent is in the middle of, '' when it is safe to exit.

    A parent's photo is accepted as a task before the handler that counts it
    ever runs, and attendance recognition, its cloud sync and its parent
    notifications all live outside that count — exiting on any of them loses
    the work outright, so each one is asked here.
    """
    reasons = []
    if _live_requests_in_flight:
        reasons.append(f"{_live_requests_in_flight} parent request(s)")
    queued = sum(1 for task in _snapshot_tasks if not task.done())
    if queued:
        reasons.append(f"{queued} queued snapshot(s)")
    attendance = attendance_engine.work_in_flight()
    if attendance:
        reasons.append(f"{attendance} attendance job(s)")
    return ", ".join(reasons)


async def _auto_update_loop() -> None:
    """Restart onto merged code by ourselves, so no one has to do it.

    The wrapper pulls origin/main every time the agent exits, so exiting on a
    quiet moment is all that is needed to run a fix the day it is merged —
    nobody has to be at the campus PC to run a restart script.
    """
    if not _AUTO_UPDATE_ENABLED:
        return
    if not _STARTED_BY_WRAPPER:
        logger.info(
            "Auto-update is off: this agent was not started by run_forever, "
            "so nothing would restart it"
        )
        return
    while True:
        try:
            await asyncio.sleep(_AUTO_UPDATE_CHECK_SECONDS)
            commit = await asyncio.to_thread(_pending_update_commit)
            if not commit:
                continue
            busy = _work_in_flight()
            if busy:
                logger.info("Update %s is waiting for %s", commit, busy)
                continue
            logger.warning(
                "Restarting onto merged code %s (running %s); the wrapper "
                "pulls it on the way back up",
                commit, _running_commit() or "unknown",
            )
            logging.shutdown()
            os._exit(0)
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.warning("Auto-update check failed: %s", _exception_text(exc))


def _exception_text(exc: BaseException) -> str:
    """Return a useful description even when str(exc) is empty."""
    detail = str(exc).strip()
    return f"{type(exc).__name__}: {detail}" if detail else type(exc).__name__


def _rtsp_cooldown_active(ip: str) -> bool:
    expires_at = _rtsp_cooldowns.get(ip, 0.0)
    if expires_at <= time.monotonic():
        _rtsp_cooldowns.pop(ip, None)
        return False
    return True


def _mark_rtsp_failure(ip: str) -> None:
    _rtsp_cooldowns[ip] = time.monotonic() + _RTSP_COOLDOWN_SECONDS
    # A stream that worked before the recorder locked itself must stop
    # vouching for the login, or every later failure goes uncounted and the
    # fallback keeps knocking on a locked account.
    if ip in _refused_credentials:
        _rtsp_credentials_worked.pop(ip, None)


def _clear_rtsp_failure(ip: str) -> None:
    _rtsp_cooldowns.pop(ip, None)


# A recorder that refuses our credentials locks the account for a while after a
# handful of rejected logins, so retrying keeps it locked and every classroom on
# it stays dark. Back off instead and take the RTSP road.
_ISAPI_AUTH_COOLDOWN_SECONDS = max(
    60.0, float(os.environ.get("ISAPI_AUTH_COOLDOWN_SECONDS", "1200"))
)
_ISAPI_TIMEOUT_COOLDOWN_SECONDS = max(
    10.0, float(os.environ.get("ISAPI_TIMEOUT_COOLDOWN_SECONDS", "120"))
)
_ISAPI_TIMEOUTS_BEFORE_BACKOFF = max(
    1, int(os.environ.get("ISAPI_TIMEOUTS_BEFORE_BACKOFF", "3"))
)
_AUTH_REJECTIONS_BEFORE_GIVING_UP = max(
    1, int(os.environ.get("ISAPI_AUTH_REJECTIONS_BEFORE_GIVING_UP", "2"))
)
# Time held back from the ISAPI attempts so the RTSP fallback still gets a turn.
_RTSP_RESERVE_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_RTSP_RESERVE_SECONDS", "8"))
)
# expires_at is None for a login refusal: retrying on a timer only re-arms the
# recorder's lockout, so that pause is held until its password changes.
_isapi_cooldowns: dict[str, tuple[float | None, str]] = {}
_isapi_consecutive_timeouts: dict[str, int] = {}
# ip -> the credentials the recorder refused, so a new password lifts the pause
_refused_credentials: dict[str, str] = {}
_auth_refused_since_ist: dict[str, str] = {}
# ip -> credentials RTSP has actually streamed with, so the fallback is only
# trusted while the recorder still accepts them
_rtsp_credentials_worked: dict[str, str] = {}
# (ip, credentials) -> RTSP attempts made since the recorder refused that login
_rtsp_attempts_while_refused: dict[tuple[str, str], int] = {}
_RTSP_ATTEMPTS_WHILE_REFUSED = max(
    1, int(os.environ.get("RTSP_ATTEMPTS_WHILE_REFUSED", "3"))
)
# A login that just served twenty classrooms cannot be wrong for the twenty
# first: when a recorder answered us moments ago, a 401 on one channel is that
# channel being unusable, not our credentials — so only that channel rests.
_AUTH_REFUSAL_TRUST_SECONDS = max(
    0.0, float(os.environ.get("ISAPI_AUTH_REFUSAL_TRUST_SECONDS", "900"))
)
_CHANNEL_AUTH_COOLDOWN_SECONDS = max(
    60.0, float(os.environ.get("ISAPI_CHANNEL_AUTH_COOLDOWN_SECONDS", "1800"))
)
# A Hikvision admin lock clears itself once nothing tries the account for a
# while, so a refused recorder is re-tried exactly once after a long silence
# instead of waiting for a human to reboot it or type a new password.
_AUTH_UNLOCK_QUIET_SECONDS = max(
    300.0, float(os.environ.get("ISAPI_AUTH_UNLOCK_QUIET_SECONDS", "1800"))
)
_AUTH_UNLOCK_MAX_QUIET_SECONDS = max(
    _AUTH_UNLOCK_QUIET_SECONDS,
    float(os.environ.get("ISAPI_AUTH_UNLOCK_MAX_QUIET_SECONDS", "14400")),
)
_AUTH_UNLOCK_CHECK_SECONDS = 60.0
# ip -> when we last presented a login to that recorder, by any path
_last_auth_attempt: dict[str, float] = {}
# ip -> earliest moment the single unlock probe may run
_auth_unlock_next_probe: dict[str, float] = {}
# ip -> how long the recorder must stay untouched before the next probe
_auth_unlock_quiet: dict[str, float] = {}
# ip -> when ISAPI last handed us a picture, so a refusal can be judged
_isapi_last_success: dict[str, float] = {}
# (ip, channel) -> when that channel may be asked for a picture again
_channel_auth_cooldowns: dict[tuple[str, int], float] = {}


class _DvrAuthRejected(Exception):
    """The recorder refused our credentials, usually a login lockout."""


def _dvr_credential_key(dvr: dict) -> str:
    """Fingerprint of a recorder's login, never the login itself."""
    return recorder_auth.credential_key(
        dvr.get("username", ""), dvr.get("password", "")
    )


def _isapi_cooldown(ip: str) -> str:
    """Why ISAPI is being skipped for this recorder, '' when it is usable."""
    entry = _isapi_cooldowns.get(ip)
    if entry is None:
        return ""
    expires_at, reason = entry
    if expires_at is None:
        return reason
    if expires_at <= time.monotonic():
        _isapi_cooldowns.pop(ip, None)
        return ""
    return reason


def _credentials_refused(dvr: dict) -> bool:
    """True while this recorder is refusing the login we still hold.

    A locked recorder re-locks itself the moment anything retries the password
    it rejected, so the only thing that lifts this is a different password.
    """
    ip = dvr.get("ip")
    refused = _refused_credentials.get(ip)
    if refused is None:
        # Another process, or this one before a restart, may already have been
        # refused; knocking again from here would only re-arm the lockout.
        if recorder_auth.is_refused(ip, _dvr_credential_key(dvr)):
            _refused_credentials[ip] = _dvr_credential_key(dvr)
            _isapi_cooldowns.setdefault(ip, (None, "credentials refused"))
            # Take the shared quiet window with it, or this process would have
            # no probe scheduled and the recorder would stay dark for good.
            _auth_unlock_quiet[ip] = recorder_auth.quiet_seconds(ip)
            due_in = recorder_auth.seconds_until_probe(ip)
            _auth_unlock_next_probe[ip] = time.monotonic() + (
                _auth_unlock_quiet[ip] if due_in is None else due_in
            )
            return True
        return False
    if refused == _dvr_credential_key(dvr):
        return True
    logger.info("%s has a new login; resuming its snapshots", ip)
    _clear_isapi_failures(ip)
    return False


def _isapi_served_recently(ip: str) -> bool:
    """Whether this recorder answered a snapshot door in the last moments."""
    last = _isapi_last_success.get(ip)
    if last is None or _LIVE_CAPTURE_BUSY_RECORDER_SECONDS <= 0:
        return False
    return (time.monotonic() - last) <= _LIVE_CAPTURE_BUSY_RECORDER_SECONDS


def _isapi_recently_worked(ip: str) -> bool:
    """Whether this recorder handed us a picture recently enough to be trusted."""
    last = _isapi_last_success.get(ip)
    if last is None:
        return False
    return (time.monotonic() - last) <= _AUTH_REFUSAL_TRUST_SECONDS


def _channel_auth_refused(ip: str, channel: int) -> bool:
    """True while this one channel is resting after refusing our login."""
    until = _channel_auth_cooldowns.get((ip, channel))
    if until is None:
        return False
    if until <= time.monotonic():
        _channel_auth_cooldowns.pop((ip, channel), None)
        return False
    return True


def _channel_doors_silent(ip: str, channel: int) -> bool:
    """True while none of this channel's snapshot doors answers at all."""
    since = _live_capture_silent_channels.get((ip, channel))
    if since is None:
        return False
    if (
        time.monotonic() - since
    ) > _LIVE_CAPTURE_SILENT_CHANNEL_TTL_SECONDS:
        _live_capture_silent_channels.pop((ip, channel), None)
        _live_capture_busy_silences.pop((ip, channel), None)
        return False
    return True


def _mark_channel_doors_silent(ip: str, channel: int) -> None:
    # A recorder that served another classroom moments ago is busy, not deaf:
    # blacklisting the channel then sends every later request down the slow
    # video road for minutes because of one crowded moment.
    if (
        (ip, channel) not in _live_capture_silent_channels
        and _isapi_served_recently(ip)
    ):
        silences = _live_capture_busy_silences.get((ip, channel), 0) + 1
        _live_capture_busy_silences[(ip, channel)] = silences
        if silences < _LIVE_CAPTURE_BUSY_FORGIVENESS:
            logger.info(
                "%s ch%d: no door answered, but the recorder served another "
                "classroom moments ago; keeping its doors in use (%d/%d)",
                ip, channel, silences, _LIVE_CAPTURE_BUSY_FORGIVENESS,
            )
            return
    if (ip, channel) not in _live_capture_silent_channels:
        logger.warning(
            "%s ch%d: no snapshot door answers; using the video stream for "
            "the next %.0fs instead of waiting on them",
            ip, channel, _LIVE_CAPTURE_SILENT_CHANNEL_TTL_SECONDS,
        )
    _live_capture_silent_channels[(ip, channel)] = time.monotonic()


def _mark_channel_auth_refused(ip: str, channel: int) -> None:
    _channel_auth_cooldowns[(ip, channel)] = (
        time.monotonic() + _CHANNEL_AUTH_COOLDOWN_SECONDS
    )
    logger.warning(
        "%s ch%d refused our login although the recorder is serving other "
        "channels; resting this channel for %.0fs instead of pausing %s",
        ip, channel, _CHANNEL_AUTH_COOLDOWN_SECONDS, ip,
    )


def _mark_isapi_auth_rejected(
    dvr: dict | str, channel: int | None = None
) -> None:
    ip = dvr["ip"] if isinstance(dvr, dict) else dvr
    if channel is not None and _isapi_recently_worked(ip):
        # Condemning the whole recorder here is what took DVR 4's 21 working
        # rooms dark because five unused channels answered 401.
        _mark_channel_auth_refused(ip, channel)
        return
    already_refused = ip in _refused_credentials
    recorder_auth.note_refusal(
        ip, _dvr_credential_key(dvr) if isinstance(dvr, dict) else ""
    )
    _isapi_cooldowns[ip] = (None, "credentials refused")
    _isapi_consecutive_timeouts.pop(ip, None)
    _refused_credentials[ip] = (
        _dvr_credential_key(dvr) if isinstance(dvr, dict) else ""
    )
    quiet = _AUTH_UNLOCK_QUIET_SECONDS
    _auth_unlock_quiet[ip] = quiet
    _auth_unlock_next_probe[ip] = time.monotonic() + quiet
    _auth_refused_since_ist.setdefault(
        ip, datetime.now(_IST).strftime("%d-%m-%Y %H:%M:%S IST")
    )
    if already_refused:
        return
    logger.error(
        "%s refused our login; every snapshot path for it is paused for "
        "%.0f minutes, because retrying only keeps the recorder locked — then "
        "it is tried once",
        ip, quiet / 60,
    )


def _rtsp_worth_trying(dvr: dict) -> bool:
    """Whether RTSP may still be tried on a recorder that refused our login.

    DVR 4 answers 401 on ISAPI yet streams over RTSP, so a refusal alone must
    not silence a recorder — but if RTSP cannot log in either, the password is
    genuinely wrong and every further attempt just re-arms the lockout.
    """
    ip = dvr.get("ip")
    key = _dvr_credential_key(dvr)
    if _rtsp_credentials_worked.get(ip) == key:
        return True
    return _rtsp_attempts_while_refused.get((ip, key), 0) < (
        _RTSP_ATTEMPTS_WHILE_REFUSED
    )


def _note_rtsp_attempt_while_refused(dvr: dict) -> None:
    ip = dvr.get("ip")
    if _refused_credentials.get(ip) is None:
        return
    key = _dvr_credential_key(dvr)
    if _rtsp_credentials_worked.get(ip) == key:
        return
    seen = _rtsp_attempts_while_refused.get((ip, key), 0) + 1
    _rtsp_attempts_while_refused[(ip, key)] = seen
    if seen == _RTSP_ATTEMPTS_WHILE_REFUSED:
        logger.error(
            "%s rejects our login on ISAPI and RTSP cannot get a frame either; "
            "leaving the recorder alone until its password changes so the "
            "lockout can clear",
            ip,
        )


def _note_rtsp_success(dvr: dict) -> None:
    ip = dvr.get("ip")
    _rtsp_credentials_worked[ip] = _dvr_credential_key(dvr)
    _rtsp_attempts_while_refused.pop((ip, _rtsp_credentials_worked[ip]), None)


def _mark_isapi_timeout(ip: str) -> None:
    count = _isapi_consecutive_timeouts.get(ip, 0) + 1
    _isapi_consecutive_timeouts[ip] = count
    if count < _ISAPI_TIMEOUTS_BEFORE_BACKOFF or _isapi_cooldown(ip):
        return
    _isapi_cooldowns[ip] = (
        time.monotonic() + _ISAPI_TIMEOUT_COOLDOWN_SECONDS,
        "not answering",
    )
    logger.warning(
        "%s failed %d captures in a row; going straight to RTSP for %.0fs",
        ip, count, _ISAPI_TIMEOUT_COOLDOWN_SECONDS,
    )


def _note_auth_attempt(ip: str) -> None:
    """Remember that a login was just presented, so the quiet window is real."""
    _last_auth_attempt[ip] = time.monotonic()
    recorder_auth.note_attempt(ip)


def _note_isapi_success(ip: str) -> None:
    """Record that the recorder just served a picture over ISAPI."""
    _isapi_last_success[ip] = time.monotonic()
    recorder_auth.note_success(ip)


def _clear_isapi_failures(ip: str) -> None:
    recorder_auth.clear(ip)
    _isapi_consecutive_timeouts.pop(ip, None)
    _refused_credentials.pop(ip, None)
    _auth_refused_since_ist.pop(ip, None)
    _auth_unlock_next_probe.pop(ip, None)
    _auth_unlock_quiet.pop(ip, None)
    if _isapi_cooldowns.pop(ip, None) is not None:
        logger.info("%s is answering ISAPI again", ip)


async def _probe_locked_recorder(dvr: dict) -> bool:
    """Try a refused recorder exactly once, after it has been left alone.

    A Hikvision admin lock expires on its own, but only while nothing presents
    the account: one quiet probe recovers the recorder without anyone at
    school rebooting it, and a failure simply lengthens the next silence.
    """
    ip = dvr.get("ip")
    port = dvr.get("port", 80)
    _note_auth_attempt(ip)
    url = f"http://{ip}:{port}/ISAPI/System/deviceInfo"
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                url,
                auth=httpx.DigestAuth(dvr["username"], dvr["password"]),
            )
        recovered = resp.status_code == 200
    except Exception as exc:
        logger.info(
            "%s did not answer the unlock check (%s); staying away",
            ip, _exception_text(exc),
        )
        recovered = False
    if recovered:
        logger.warning(
            "%s accepted our login again after the lockout expired; resuming "
            "its cameras", ip,
        )
        _note_isapi_success(ip)
        for key in [k for k in _rtsp_attempts_while_refused if k[0] == ip]:
            _rtsp_attempts_while_refused.pop(key, None)
        _clear_isapi_failures(ip)
        return True
    quiet = min(
        _AUTH_UNLOCK_MAX_QUIET_SECONDS,
        _auth_unlock_quiet.get(ip, _AUTH_UNLOCK_QUIET_SECONDS) * 2,
    )
    recorder_auth.note_probe_failed(ip)
    _auth_unlock_quiet[ip] = quiet
    _auth_unlock_next_probe[ip] = time.monotonic() + quiet
    logger.info(
        "%s still refuses our login; leaving it untouched for %.0f more "
        "minutes", ip, quiet / 60,
    )
    return False


async def _unlock_refused_recorders() -> None:
    """One pass of the unlock watch over every recorder that refused us."""
    for dvr in config.get("dvrs", []) or []:
        ip = dvr.get("ip")
        if not ip or not _credentials_refused(dvr):
            continue
        due_at = _auth_unlock_next_probe.get(ip)
        if due_at is None or due_at > time.monotonic():
            continue
        quiet = _auth_unlock_quiet.get(ip, _AUTH_UNLOCK_QUIET_SECONDS)
        touched = _last_auth_attempt.get(ip)
        if touched is not None and (time.monotonic() - touched) < quiet:
            # Something presented the login again (an RTSP fallback, say), so
            # the lock was re-armed and the silence has to start over.
            _auth_unlock_next_probe[ip] = touched + quiet
            continue
        # The gate counter and the mood watcher log in from their own
        # processes: probing while they are still knocking would only find the
        # lock they keep re-arming.
        elsewhere = recorder_auth.seconds_since_attempt(ip)
        if elsewhere is not None and elsewhere < quiet:
            _auth_unlock_next_probe[ip] = time.monotonic() + (quiet - elsewhere)
            continue
        await _probe_locked_recorder(dvr)


async def _unlock_watch_loop() -> None:
    while True:
        try:
            await asyncio.sleep(_AUTH_UNLOCK_CHECK_SECONDS)
            await _unlock_refused_recorders()
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.warning("Unlock watch failed (non-fatal): %s", exc)


def dvr_snapshot_health() -> list[dict]:
    """Recorders whose direct snapshot path is currently being avoided."""
    health = []
    for ip, (expires_at, reason) in sorted(_isapi_cooldowns.items()):
        if expires_at is None:
            probe_at = _auth_unlock_next_probe.get(ip)
            health.append({
                "ip": ip,
                "reason": reason,
                "seconds_remaining": None,
                "held_until_password_change": probe_at is None,
                "retry_in_seconds": (
                    None if probe_at is None
                    else round(max(0.0, probe_at - time.monotonic()), 1)
                ),
                "since_ist": _auth_refused_since_ist.get(ip, ""),
            })
        elif expires_at > time.monotonic():
            health.append({
                "ip": ip,
                "reason": reason,
                "seconds_remaining": round(
                    max(0.0, expires_at - time.monotonic()), 1
                ),
            })
    return health


# The first frames a recorder's video stream decodes are usually grey filler
# while the decoder waits for a keyframe, so a picture is only worth sending
# once it actually carries detail.
_RTSP_MIN_FRAME_STDDEV = max(
    0.0, float(os.environ.get("RTSP_MIN_FRAME_STDDEV", "6"))
)
_RTSP_MAX_FRAMES_READ = max(
    1, int(os.environ.get("RTSP_MAX_FRAMES_READ", "30"))
)
_RTSP_FRAME_SEARCH_SECONDS = max(
    0.5, float(os.environ.get("RTSP_FRAME_SEARCH_SECONDS", "5"))
)


def _frame_carries_detail(frame) -> bool:
    """False for the flat grey frame a decoder emits before its first keyframe."""
    try:
        return float(frame.std()) >= _RTSP_MIN_FRAME_STDDEV
    except Exception:
        return True


def _read_detailed_frame(cap, ip: str, channel: int):
    """Read past the decoder's grey filler frames to a real classroom picture."""
    deadline = time.monotonic() + _RTSP_FRAME_SEARCH_SECONDS
    blank = 0
    for _ in range(_RTSP_MAX_FRAMES_READ):
        ret, frame = cap.read()
        if not ret or frame is None:
            break
        if _frame_carries_detail(frame):
            if blank:
                logger.info(
                    "[RTSP] %s ch%d: skipped %d blank frame(s) before a real "
                    "picture", ip, channel, blank,
                )
            return frame
        blank += 1
        if time.monotonic() >= deadline:
            break
    if blank:
        logger.warning(
            "[RTSP] %s ch%d: every one of %d frame(s) was blank, refusing to "
            "send a grey photo", ip, channel, blank,
        )
    return None


def _warn_rtsp_timeouts_unsupported() -> None:
    global _rtsp_timeout_warning_logged
    with _rtsp_timeout_warning_lock:
        if _rtsp_timeout_warning_logged:
            return
        logger.warning(
            "[RTSP] OpenCV build does not support timeout parameters; "
            "using unbounded RTSP open/read calls"
        )
        _rtsp_timeout_warning_logged = True


def _capture_frame_rtsp(ip: str, port: int, user: str, pwd: str,
                        channel: int) -> bytes | None:
    """Capture a single JPEG frame via RTSP using OpenCV.

    This is a synchronous fallback for DVRs where ISAPI returns 401 but RTSP works.
    """
    try:
        import cv2
    except ImportError:
        logger.warning("[RTSP] cv2 not available for RTSP fallback")
        return None

    stream_channel = channel * 100 + 1
    # URL-encode the @ in password
    safe_pwd = pwd.replace("@", "%40")
    rtsp_url = f"rtsp://{user}:{safe_pwd}@{ip}:{port}/Streaming/Channels/{stream_channel}"
    logger.debug("[RTSP] Attempting capture from %s ch%d", ip, channel)

    cap = None
    try:
        try:
            cap = cv2.VideoCapture()
            timeout_params = []
            for property_name in (
                "CAP_PROP_OPEN_TIMEOUT_MSEC",
                "CAP_PROP_READ_TIMEOUT_MSEC",
            ):
                property_id = getattr(cv2, property_name, None)
                if property_id is not None:
                    timeout_params.extend(
                        (property_id, _RTSP_CAPTURE_TIMEOUT_MILLISECONDS)
                    )
            cap.open(rtsp_url, cv2.CAP_FFMPEG, timeout_params)
        except Exception:
            _warn_rtsp_timeouts_unsupported()
            if cap is not None:
                cap.release()
            cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
        if not cap.isOpened():
            logger.warning("[RTSP] Failed to open %s ch%d", ip, channel)
            return None
        frame = _read_detailed_frame(cap, ip, channel)
        if frame is None:
            logger.warning(
                "[RTSP] No usable frame from %s ch%d", ip, channel
            )
            return None
        ok, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 90])
        if ok:
            logger.info("[RTSP] Captured frame from %s ch%d (%d bytes)",
                        ip, channel, len(buf))
            return buf.tobytes()
        return None
    except Exception as e:
        logger.error(
            "[RTSP] Error capturing from %s ch%d: %s",
            ip,
            channel,
            _exception_text(e),
        )
        return None
    finally:
        if cap is not None:
            cap.release()


async def _capture_snapshot_rtsp(
    dvr: dict, channel: int, background: bool = False
) -> bytes | None:
    """Async wrapper around synchronous RTSP capture.

    Classroom scanning gets its own slots so a parent waiting for a live photo
    never queues behind the scanner's sweep of every camera.
    """
    loop = asyncio.get_running_loop()
    # RTSP presents the same account, so an unlock probe must not run while a
    # stream is re-arming the lock — unless RTSP is logging in fine, which
    # proves the account is not locked at all.
    # Every stream presents the account, so each one restarts the silence a
    # locked recorder needs — even where RTSP logged in fine a moment ago.
    _note_auth_attempt(dvr["ip"])
    semaphore = (
        _rtsp_background_semaphore if background else _rtsp_capture_semaphore
    )
    await semaphore.acquire()
    try:
        future = _rtsp_capture_executor.submit(
            _capture_frame_rtsp,
            dvr["ip"],
            554,
            dvr["username"],
            dvr["password"],
            channel,
        )
    except Exception:
        semaphore.release()
        raise

    def release_slot(_future: concurrent.futures.Future) -> None:
        if not loop.is_closed():
            loop.call_soon_threadsafe(semaphore.release)

    future.add_done_callback(release_slot)
    return await asyncio.wrap_future(future)


_SNAPSHOT_RETRIES = max(1, int(os.environ.get("SNAPSHOT_RETRIES", "3")))
_SNAPSHOT_RETRY_BACKOFF_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_RETRY_BACKOFF_SECONDS", "0.4"))
)
_SNAPSHOT_CONNECT_TIMEOUT_SECONDS = max(
    1.0, float(os.environ.get("SNAPSHOT_CONNECT_TIMEOUT_SECONDS", "5"))
)
_SNAPSHOT_HTTP_TIMEOUT_SECONDS = max(
    _SNAPSHOT_CONNECT_TIMEOUT_SECONDS,
    float(os.environ.get("SNAPSHOT_HTTP_TIMEOUT_SECONDS", "10")),
)
_SNAPSHOT_BACKGROUND_RETRIES = max(
    1, int(os.environ.get("SNAPSHOT_BACKGROUND_RETRIES", "1"))
)
_SNAPSHOT_BACKGROUND_CAMERA_TIMEOUT_SECONDS = max(
    1.0, float(os.environ.get("SNAPSHOT_BACKGROUND_CAMERA_TIMEOUT_SECONDS", "8"))
)
_SNAPSHOT_BACKGROUND_HTTP_TIMEOUT_SECONDS = max(
    1.0, float(os.environ.get("SNAPSHOT_BACKGROUND_HTTP_TIMEOUT_SECONDS", "4"))
)
_SNAPSHOT_BACKGROUND_CONNECT_TIMEOUT_SECONDS = max(
    1.0, float(os.environ.get("SNAPSHOT_BACKGROUND_CONNECT_TIMEOUT_SECONDS", "3"))
)
_SNAPSHOT_BACKGROUND_RETRY_BACKOFF_SECONDS = max(
    0.0, float(os.environ.get("SNAPSHOT_BACKGROUND_RETRY_BACKOFF_SECONDS", "0.1"))
)
_RTSP_CAPTURE_CONCURRENCY = max(
    1, int(os.environ.get("RTSP_CAPTURE_CONCURRENCY", "4"))
)
_RTSP_BACKGROUND_CONCURRENCY = max(
    1, int(os.environ.get("RTSP_BACKGROUND_CONCURRENCY", "2"))
)
_RTSP_CAPTURE_TIMEOUT_MILLISECONDS = max(
    1000, int(os.environ.get("RTSP_CAPTURE_TIMEOUT_MILLISECONDS", "3000"))
)
_rtsp_capture_executor = concurrent.futures.ThreadPoolExecutor(
    max_workers=_RTSP_CAPTURE_CONCURRENCY + _RTSP_BACKGROUND_CONCURRENCY,
    thread_name_prefix="ppis-rtsp",
)
_rtsp_capture_semaphore = asyncio.Semaphore(_RTSP_CAPTURE_CONCURRENCY)
_rtsp_background_semaphore = asyncio.Semaphore(_RTSP_BACKGROUND_CONCURRENCY)
_SNAPSHOT_LIVE_REQUEST_BUDGET_SECONDS = max(
    1.0, float(os.environ.get("SNAPSHOT_LIVE_REQUEST_BUDGET_SECONDS", "50"))
)


async def _get_live_dvr_client(
    dvr: dict, timeout: httpx.Timeout
) -> httpx.AsyncClient:
    ip = dvr["ip"]
    async with _live_dvr_client_lock:
        client = _live_dvr_clients.get(ip)
        if client is None or getattr(client, "is_closed", False):
            _live_dvr_clients[ip] = httpx.AsyncClient(
                timeout=timeout,
                limits=httpx.Limits(
                    max_connections=_DVR_CAPTURE_LIMIT,
                    max_keepalive_connections=_DVR_CAPTURE_LIMIT,
                ),
            )
        return _live_dvr_clients[ip]


async def _evict_live_dvr_client(
    ip: str, client: httpx.AsyncClient
) -> None:
    async with _live_dvr_client_lock:
        if _live_dvr_clients.get(ip) is not client:
            return
        _live_dvr_clients.pop(ip, None)


def _response_round_trips(response: httpx.Response) -> int:
    history = getattr(response, "history", ())
    return 1 + len(history)


def _live_client_should_evict(exc: BaseException) -> bool:
    return isinstance(
        exc,
        (
            httpx.ConnectError,
            httpx.ConnectTimeout,
            httpx.RemoteProtocolError,
            httpx.ReadError,
            httpx.WriteError,
            ConnectionResetError,
        ),
    )


async def _capture_snapshot_once(
    dvr: dict,
    channel: int,
    client: httpx.AsyncClient,
    metrics: dict[str, object] | None = None,
    door_budget_seconds: float | None = None,
) -> bytes | None:
    ip = dvr["ip"]
    port = dvr.get("port", 80)
    user = dvr["username"]
    pwd = dvr["password"]
    stream_channel = channel * 100 + 1
    url = f"http://{ip}:{port}/ISAPI/Streaming/channels/{stream_channel}/picture"
    full_size = url
    if _LIVE_SNAPSHOT_WIDTH and _LIVE_SNAPSHOT_HEIGHT:
        full_size = (
            f"{url}?snapShotImageType=JPEG"
            f"&videoResolutionWidth={_LIVE_SNAPSHOT_WIDTH}"
            f"&videoResolutionHeight={_LIVE_SNAPSHOT_HEIGHT}"
        )
    urls = [
        full_size,
        url,
        f"http://{ip}:{port}/ISAPI/Streaming/channels/{channel * 100 + 2}/picture",
        f"http://{ip}:{port}/Streaming/channels/{stream_channel}/picture",
    ]
    digest_auth = _digest_auth(ip, user, pwd)
    candidates = []
    key = (ip, channel)
    remembered = _live_capture_preferences.get(key)
    if remembered is not None and remembered[1] != 0:
        age = time.monotonic() - _live_capture_preference_age.get(key, 0.0)
        if age > _LIVE_CAPTURE_FALLBACK_TTL_SECONDS:
            _live_capture_preferences.pop(key, None)
            _live_capture_best_pixels.pop(key, None)
            remembered = None
    if remembered is not None:
        candidates.append(remembered)
    candidates.extend(
        (scheme, variant)
        for variant in range(len(urls))
        for scheme in ("digest", "basic")
    )
    candidates = [
        candidate for candidate in candidates
        if candidate[1] <= 1 or candidate[0] == "digest"
    ]
    slow_doors = _live_capture_slow_doors.get(key)
    if slow_doors:
        fresh = {
            variant: hung_at for variant, hung_at in slow_doors.items()
            if time.monotonic() - hung_at <= _LIVE_CAPTURE_SLOW_DOOR_TTL_SECONDS
        }
        if fresh:
            _live_capture_slow_doors[key] = fresh
        else:
            _live_capture_slow_doors.pop(key, None)
        slow_doors = fresh
    if slow_doors:
        # A door that hung goes last: the parent's photo is behind whichever
        # door still answers, not behind the one that never does.
        candidates.sort(key=lambda candidate: candidate[1] in slow_doors)
    seen: set[tuple[str, int]] = set()
    auth_rejections = 0
    usable_replies = 0
    last_rejection = 0
    wanted_pixels = _LIVE_SNAPSHOT_WIDTH * _LIVE_SNAPSHOT_HEIGHT
    known_best = _live_capture_best_pixels.get(key, 0)
    best: tuple[str, int, bytes, int] | None = None
    pictures = 0
    sized_variants: set[int] = set()
    started = time.monotonic()

    def keep(scheme: str, variant: int, picture: bytes, pixels: int) -> bytes:
        if _live_capture_preferences.get(key) != (scheme, variant):
            # Age the choice from when it was made, so a busy camera stuck
            # on a fallback still retries the full-size stream.
            _live_capture_preference_age[key] = time.monotonic()
        _live_capture_preferences[key] = (scheme, variant)
        _live_capture_best_pixels[key] = max(known_best, pixels)
        if (
            pixels
            and wanted_pixels
            and pixels < wanted_pixels
            and key not in _live_capture_size_logged
        ):
            _live_capture_size_logged.add(key)
            logger.warning(
                "%s ch%d only serves %d pixels, less than the %d asked for: "
                "the parent's photo will look soft",
                ip, channel, pixels, wanted_pixels,
            )
        return picture

    silent_variants: set[int] = set()
    for scheme, variant in candidates:
        if (scheme, variant) in seen or variant in sized_variants:
            continue
        if variant in silent_variants:
            # A URL that never answers will not start answering for the other
            # authentication scheme; spend the attempt on the next URL instead.
            continue
        seen.add((scheme, variant))
        auth = digest_auth if scheme == "digest" else httpx.BasicAuth(user, pwd)
        door_budget = (
            _LIVE_CAPTURE_DOOR_TIMEOUT_SECONDS
            if door_budget_seconds is None else door_budget_seconds
        )
        if best is not None:
            # Probing for something sharper must never cost the picture we have.
            door_budget = min(
                door_budget,
                _LIVE_CAPTURE_PROBE_BUDGET_SECONDS - (time.monotonic() - started),
            )
        try:
            response = await asyncio.wait_for(
                client.get(urls[variant], auth=auth),
                timeout=max(0.1, door_budget),
            )
        except (asyncio.TimeoutError, httpx.ReadTimeout):
            if best is not None:
                return keep(*best)
            _live_capture_slow_doors.setdefault(key, {})[variant] = time.monotonic()
            silent_variants.add(variant)
            if metrics is not None:
                metrics["door_timeouts"] = metrics.get("door_timeouts", 0) + 1
            logger.warning(
                "%s ch%d: no answer within %.1fs from %s, trying the next door",
                ip, channel, door_budget, urls[variant],
            )
            continue
        except httpx.HTTPError:
            if best is not None:
                return keep(*best)
            raise
        if metrics is not None:
            metrics["door"] = variant
            metrics["answered"] = True
            metrics["round_trips"] = (
                metrics.get("round_trips", 0) + _response_round_trips(response)
            )
        if response.status_code in (401, 403):
            auth_rejections += 1
            last_rejection = response.status_code
            continue
        if response.status_code < 400:
            # Only a door that could have served a picture proves the recorder
            # still accepts us; a 404 door must not hide a login lockout.
            usable_replies += 1
        if (
            response.status_code == 200
            and response.headers.get("content-type", "").startswith("image")
        ):
            pixels = _jpeg_pixels(response.content)
            sized_variants.add(variant)
            if key in _live_capture_slow_doors:
                _live_capture_slow_doors[key].pop(variant, None)
            if best is None or pixels > best[3]:
                best = (scheme, variant, response.content, pixels)
            pictures += 1
            if (
                not pixels
                or pixels >= wanted_pixels
                or (known_best and pixels >= known_best)
                or pictures >= _LIVE_CAPTURE_PROBE_PICTURES
                or time.monotonic() - started >= _LIVE_CAPTURE_PROBE_BUDGET_SECONDS
            ):
                # As sharp as this channel is known to get, so take it.
                return keep(*best)
            # Undersized on a channel we have never sized up: the sub-stream may
            # have answered where the main stream should have, so try one more
            # door before settling for a soft picture.
    if best is not None:
        return keep(*best)
    if (
        not usable_replies
        and auth_rejections >= _AUTH_REJECTIONS_BEFORE_GIVING_UP
    ):
        # Every door was locked, so the credentials are wrong for this
        # recorder; hammering it only deepens its login lockout.
        raise _DvrAuthRejected(
            f"{ip} rejected our credentials (HTTP {last_rejection})"
        )
    return None


async def capture_snapshot(
    dvr: dict,
    channel: int,
    *,
    background: bool = False,
    classroom: str = "",
    deadline: float | None = None,
) -> bytes | None:
    """Capture a live JPEG, sharing one capture between parents asking together.

    Several parents of one class used to queue a separate capture each, which is
    what turned a crowded minute into a half-minute wait: the recorder was
    answering, just for everybody at once. The classroom scanner keeps its own
    capture — recognising faces needs the frame of its own sweep.
    """
    if background or not _LIVE_CAPTURE_SHARING:
        return await _capture_snapshot_now(
            dvr,
            channel,
            background=background,
            classroom=classroom,
            deadline=deadline,
        )
    key = (dvr["ip"], channel)
    report = _live_capture_report.get()
    task = _live_capture_in_flight.get(key)
    mine = task is None or task.done()
    if mine:
        task = asyncio.create_task(
            _capture_snapshot_now(
                dvr,
                channel,
                background=False,
                classroom=classroom,
                deadline=deadline,
            )
        )
        _live_capture_in_flight[key] = task
    else:
        logger.info(
            "%s ch%d: waiting on the capture already running for %s",
            key[0], channel, classroom or _live_request_classroom.get() or "-",
        )
    try:
        # Shielded, so a caller whose own request times out does not cancel the
        # capture the other waiting parents are relying on.
        picture = await asyncio.shield(task)
    finally:
        # Only once the capture is over: the first parent's request can time out
        # while it is still running, and the parents behind them must join that
        # capture rather than start another one on the same camera.
        if task.done() and _live_capture_in_flight.get(key) is task:
            _live_capture_in_flight.pop(key, None)
    if not mine and report is not None and not report:
        report.update({
            "seconds": 0.0,
            "shared": True,
            "recorder": key[0],
            "channel": channel,
            "outcome": "success" if picture else "failed",
        })
    return picture


async def _capture_snapshot_now(
    dvr: dict,
    channel: int,
    *,
    background: bool = False,
    classroom: str = "",
    deadline: float | None = None,
) -> bytes | None:
    """Capture a JPEG snapshot from a Hikvision NVR via ISAPI.

    Hikvision DS-9664NI-ST / DS-7632NXI-K2 supports:
      GET /ISAPI/Streaming/channels/{channel}01/picture
    where channel is 1-based (1..64 per NVR).

    Returns JPEG bytes or None on failure.
    Sets _last_capture_error with diagnostic details.
    """
    global _last_capture_error
    ip = dvr["ip"]
    stream_channel = channel * 100 + 1
    capture_started = time.monotonic()
    logger.info(
        "Capturing snapshot from %s channel %d (stream %d)",
        ip,
        channel,
        stream_channel,
    )

    limiter_wait_started = time.monotonic()
    limiter = await _acquire_dvr_capture(ip, background)
    limiter_wait = time.monotonic() - limiter_wait_started
    metrics: dict[str, object] = {
        "round_trips": 0,
        "attempt_elapsed": [],
        "rtsp": False,
        "outcome": "failed",
        "exception": "",
        "door_timeouts": 0,
        "door": -1,
        "answered": False,
    }
    try:
        retries = _SNAPSHOT_BACKGROUND_RETRIES if background else _SNAPSHOT_RETRIES
        camera_timeout = (
            _SNAPSHOT_BACKGROUND_CAMERA_TIMEOUT_SECONDS
            if background else _SNAPSHOT_CAMERA_TIMEOUT_SECONDS
        )
        http_timeout = (
            _SNAPSHOT_BACKGROUND_HTTP_TIMEOUT_SECONDS
            if background else _SNAPSHOT_HTTP_TIMEOUT_SECONDS
        )
        connect_timeout = (
            _SNAPSHOT_BACKGROUND_CONNECT_TIMEOUT_SECONDS
            if background else _SNAPSHOT_CONNECT_TIMEOUT_SECONDS
        )
        request_deadline = deadline if deadline is not None else _live_request_deadline.get()
        capture_deadline = capture_started + camera_timeout
        if not background and request_deadline is not None:
            capture_deadline = min(capture_deadline, request_deadline)
        timeout = httpx.Timeout(
            http_timeout,
            connect=connect_timeout,
        )
        client = await _get_live_dvr_client(dvr, timeout) if not background else None
        if background:
            client = httpx.AsyncClient(timeout=timeout)
        # Keep part of the budget for the RTSP fallback, otherwise a recorder
        # that never answers eats the whole request and the parent gets nothing.
        isapi_deadline = capture_deadline
        if not background:
            isapi_deadline -= min(
                _RTSP_RESERVE_SECONDS, max(0.0, camera_timeout * 0.5)
            )
        # A recorder that refused this password re-arms its own lockout on every
        # further login attempt, so once RTSP has been shown to fail with the
        # same credentials nothing is tried again until the password changes.
        if _credentials_refused(dvr) and not _rtsp_worth_trying(dvr):
            _last_capture_error = (
                f"{ip} ch{channel}: recorder refused our login; letting its "
                f"lockout expire before one quiet retry (refused since "
                f"{_auth_refused_since_ist.get(ip, 'today')})"
            )
            metrics["outcome"] = "credentials_refused"
            logger.warning(
                "%s ch%d: skipped, recorder is refusing our login and RTSP "
                "cannot log in either", ip, channel,
            )
            return None
        cooldown_reason = _isapi_cooldown(ip)
        # While a recorder is on cooldown the RTSP road is the one worth taking,
        # unless that has just failed too — then one direct try beats no picture.
        skip_isapi = bool(cooldown_reason) and not _rtsp_cooldown_active(ip)
        # ...but never against a recorder that refused this login: one more
        # rejected attempt is what keeps the lockout alive.
        if cooldown_reason == "credentials refused":
            skip_isapi = True
        if (
            not cooldown_reason
            and not _rtsp_cooldown_active(ip)
            and _channel_doors_silent(ip, channel)
        ):
            # Every door of this channel was silent a moment ago; waiting on
            # them again only delays the picture the video stream can give.
            skip_isapi = True
            _last_capture_error = (
                f"{ip} ch{channel}: no snapshot door answers"
            )
            logger.info(
                "%s ch%d: snapshot doors are silent, going straight to RTSP",
                ip, channel,
            )
        elif not cooldown_reason and _channel_auth_refused(ip, channel):
            # This one channel refuses us while its recorder serves others.
            skip_isapi = True
            _last_capture_error = (
                f"{ip} ch{channel}: channel refused our login"
            )
            logger.info(
                "%s ch%d: channel is resting after refusing our login, "
                "using RTSP", ip, channel,
            )
        elif cooldown_reason:
            logger.info(
                "%s ch%d: recorder is %s, %s",
                ip, channel, cooldown_reason,
                "using RTSP" if skip_isapi else "trying ISAPI once anyway",
            )
            _last_capture_error = f"{ip} ch{channel}: ISAPI paused ({cooldown_reason})"
        isapi_timed_out = False
        if skip_isapi:
            attempts = 0
        else:
            attempts = 1 if cooldown_reason else retries
        try:
            for attempt in range(attempts):
                remaining = isapi_deadline - time.monotonic()
                if remaining <= 0:
                    break
                attempt_started = time.monotonic()
                _note_auth_attempt(ip)
                try:
                    snapshot = await asyncio.wait_for(
                        _capture_snapshot_once(
                            dvr,
                            channel,
                            client,
                            metrics,
                            # Two doors must fit inside one attempt, or a silent
                            # first door still costs every later one — the
                            # background sweep's attempt is shorter than the
                            # default door budget.
                            door_budget_seconds=min(
                                _LIVE_CAPTURE_DOOR_TIMEOUT_SECONDS,
                                max(1.0, min(http_timeout, remaining) / 2),
                            ),
                        ),
                        timeout=min(http_timeout, remaining),
                    )
                    metrics["attempt_elapsed"].append(
                        round(time.monotonic() - attempt_started, 3)
                    )
                    if snapshot:
                        logger.info(
                            "Snapshot captured: %d bytes from %s ch%d in %.2fs",
                            len(snapshot), ip, channel,
                            time.monotonic() - capture_started,
                        )
                        _last_capture_error = ""
                        metrics["outcome"] = "success"
                        _isapi_last_success[ip] = time.monotonic()
                        _channel_auth_cooldowns.pop((ip, channel), None)
                        _live_capture_silent_channels.pop((ip, channel), None)
                        _live_capture_busy_silences.pop((ip, channel), None)
                        _clear_isapi_failures(ip)
                        return snapshot
                    _last_capture_error = f"{ip} ch{channel}: ISAPI capture failed"
                    if metrics["door_timeouts"]:
                        isapi_timed_out = True
                    if metrics["door_timeouts"] and not metrics["answered"]:
                        # Not one door on this channel replied: further attempts
                        # knock on the same silent doors and cost the parent the
                        # time the video fallback needs.
                        _mark_channel_doors_silent(ip, channel)
                        break
                    if attempt < attempts - 1:
                        await asyncio.sleep(min(
                            _SNAPSHOT_RETRY_BACKOFF_SECONDS if not background
                            else _SNAPSHOT_BACKGROUND_RETRY_BACKOFF_SECONDS,
                            max(0.0, isapi_deadline - time.monotonic()),
                        ))
                except _DvrAuthRejected as exc:
                    metrics["attempt_elapsed"].append(
                        round(time.monotonic() - attempt_started, 3)
                    )
                    metrics["exception"] = type(exc).__name__
                    _last_capture_error = f"{ip} ch{channel}: {exc}"
                    _mark_isapi_auth_rejected(dvr, channel)
                    break
                except (httpx.ConnectError, httpx.ConnectTimeout, httpx.ReadTimeout) as exc:
                    metrics["attempt_elapsed"].append(
                        round(time.monotonic() - attempt_started, 3)
                    )
                    metrics["exception"] = type(exc).__name__
                    isapi_timed_out = True
                    if not background and _live_client_should_evict(exc):
                        await _evict_live_dvr_client(ip, client)
                    _last_capture_error = f"{ip} ch{channel}: {_exception_text(exc)}"
                    logger.warning(
                        "Snapshot attempt %d/%d failed from %s ch%d: %s",
                        attempt + 1, attempts, ip, channel, _exception_text(exc),
                    )
                    if isinstance(exc, httpx.ReadTimeout) and not metrics["answered"]:
                        _mark_channel_doors_silent(ip, channel)
                        break
                    if attempt < attempts - 1:
                        await asyncio.sleep(min(
                            _SNAPSHOT_RETRY_BACKOFF_SECONDS if not background
                            else _SNAPSHOT_BACKGROUND_RETRY_BACKOFF_SECONDS,
                            max(0.0, isapi_deadline - time.monotonic()),
                        ))
                except asyncio.TimeoutError as exc:
                    metrics["attempt_elapsed"].append(
                        round(time.monotonic() - attempt_started, 3)
                    )
                    metrics["exception"] = type(exc).__name__
                    isapi_timed_out = True
                    _last_capture_error = f"{ip} ch{channel}: {_exception_text(exc)}"
                    logger.warning(
                        "Snapshot attempt %d/%d timed out from %s ch%d",
                        attempt + 1, attempts, ip, channel,
                    )
                    if not metrics["answered"]:
                        _mark_channel_doors_silent(ip, channel)
                        break
                    if attempt < attempts - 1:
                        await asyncio.sleep(min(
                            _SNAPSHOT_RETRY_BACKOFF_SECONDS if not background
                            else _SNAPSHOT_BACKGROUND_RETRY_BACKOFF_SECONDS,
                            max(0.0, isapi_deadline - time.monotonic()),
                        ))
                except Exception as exc:
                    metrics["attempt_elapsed"].append(
                        round(time.monotonic() - attempt_started, 3)
                    )
                    metrics["exception"] = type(exc).__name__
                    if not background and _live_client_should_evict(exc):
                        await _evict_live_dvr_client(ip, client)
                    _last_capture_error = f"{ip} ch{channel}: {_exception_text(exc)}"
                    logger.error(
                        "Snapshot error from %s ch%d: %s",
                        ip,
                        channel,
                        _exception_text(exc),
                    )
                    if attempt < attempts - 1:
                        await asyncio.sleep(min(
                            _SNAPSHOT_RETRY_BACKOFF_SECONDS if not background
                            else _SNAPSHOT_BACKGROUND_RETRY_BACKOFF_SECONDS,
                            max(0.0, isapi_deadline - time.monotonic()),
                        ))
            # Once per request, not once per retry: three retries inside one
            # slow request must not look like three failing requests.
            if isapi_timed_out:
                _mark_isapi_timeout(ip)

            remaining = capture_deadline - time.monotonic()
            if (
                remaining > 0
                and (
                    not background
                    or ip in _RTSP_FALLBACK_IPS
                    or _isapi_cooldown(ip)
                )
                and not _rtsp_cooldown_active(ip)
            ):
                logger.info("ISAPI exhausted for %s ch%d, trying RTSP fallback", ip, channel)
                _note_rtsp_attempt_while_refused(dvr)
                try:
                    rtsp_frame = await asyncio.wait_for(
                        _capture_snapshot_rtsp(
                            dvr, channel, background=background
                        ),
                        timeout=remaining,
                    )
                    if rtsp_frame:
                        metrics["rtsp"] = True
                        metrics["outcome"] = "success"
                        _clear_rtsp_failure(ip)
                        _note_rtsp_success(dvr)
                        _last_capture_error = ""
                        return rtsp_frame
                    _mark_rtsp_failure(ip)
                    metrics["rtsp"] = True
                except asyncio.TimeoutError:
                    _mark_rtsp_failure(ip)
                    metrics["rtsp"] = True
                    logger.warning(
                        "RTSP fallback timed out from %s ch%d",
                        ip,
                        channel,
                    )
                except Exception as exc:
                    _mark_rtsp_failure(ip)
                    metrics["rtsp"] = True
                    metrics["exception"] = type(exc).__name__
                    logger.warning(
                        "RTSP fallback failed from %s ch%d: %s",
                        ip,
                        channel,
                        _exception_text(exc),
                    )
            elif remaining > 0 and not background and _rtsp_cooldown_active(ip):
                logger.info(
                    "Skipping RTSP fallback for %s ch%d during failure cooldown",
                    ip,
                    channel,
                )
            return None
        finally:
            if background and client is not None:
                close = getattr(client, "aclose", None)
                if close is not None:
                    await close()
    finally:
        await limiter.release(not background)
        if not background:
            elapsed = time.monotonic() - capture_started
            logger.info(
                "[LIVE] snapshot classroom=%s ip=%s channel=%d elapsed=%.2fs "
                "limiter_wait=%.3fs http_round_trips=%d attempt_elapsed=%s "
                "door=%s door_timeouts=%d rtsp=%s outcome=%s exception=%s",
                classroom or _live_request_classroom.get() or "-",
                ip,
                channel,
                elapsed,
                limiter_wait,
                metrics["round_trips"],
                metrics["attempt_elapsed"],
                metrics["door"],
                metrics["door_timeouts"],
                metrics["rtsp"],
                metrics["outcome"],
                metrics["exception"] or "-",
            )
            sink = _live_capture_report.get()
            if sink is not None:
                # The campus PC's log is not reachable from the cloud, so the
                # timing travels with the photo instead.
                sink.update({
                    "seconds": round(elapsed, 2),
                    "slot_wait_seconds": round(limiter_wait, 2),
                    "attempt_seconds": list(metrics["attempt_elapsed"]),
                    "door": metrics["door"],
                    "door_timeouts": metrics["door_timeouts"],
                    "rtsp": bool(metrics["rtsp"]),
                    "recorder": ip,
                    "channel": channel,
                    "outcome": str(metrics["outcome"]),
                    "exception": str(metrics["exception"]),
                })


async def test_dvr_connection(dvr: dict) -> dict:
    """Test connection to a DVR and return status info."""
    ip = dvr["ip"]
    port = dvr.get("port", 80)
    user = dvr["username"]
    pwd = dvr["password"]

    # Testing a recorder that already refused this password is what kept DVR 2
    # locked: every test re-armed the lockout, so answer from what we know.
    if _credentials_refused(dvr):
        return {
            "status": "auth_failed",
            "ip": ip,
            "error": (
                "Recorder refused this username/password at "
                f"{_auth_refused_since_ist.get(ip, 'an earlier check')}; not "
                "retrying it now, because every attempt keeps the recorder "
                "locked. It is tried again automatically once the recorder "
                "has been left alone long enough for the lockout to expire."
            ),
            "held_until_password_change": (
                _auth_unlock_next_probe.get(ip) is None
            ),
            "retry_in_seconds": (
                None if _auth_unlock_next_probe.get(ip) is None
                else round(
                    max(0.0, _auth_unlock_next_probe[ip] - time.monotonic()), 1
                )
            ),
        }

    url = f"http://{ip}:{port}/ISAPI/System/deviceInfo"
    _note_auth_attempt(ip)
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(url, auth=httpx.DigestAuth(user, pwd))
            if resp.status_code == 200:
                _isapi_last_success[ip] = time.monotonic()
                _clear_isapi_failures(ip)
                return {"status": "connected", "ip": ip, "response": resp.text[:500]}
            elif resp.status_code == 401:
                _mark_isapi_auth_rejected(dvr)
                return {"status": "auth_failed", "ip": ip, "error": "Invalid username/password"}
            else:
                return {"status": "error", "ip": ip, "error": f"HTTP {resp.status_code}"}
    except httpx.ConnectError:
        return {"status": "unreachable", "ip": ip, "error": "Cannot connect — DVR may be offline or IP is wrong"}
    except Exception as e:
        return {"status": "error", "ip": ip, "error": str(e)}


# Channel names read from each DVR, so a classroom whose second camera was
# never entered in the mapping is still captured: {dvr_index: {name: channel}}
_discovered_dvr_channels: dict[int, dict[str, int]] = {}


def _camera_name_base(description: str) -> str:
    """'G8B C1' -> 'G8B'; the part identifying the room, not the angle."""
    return re.sub(
        r"\s*C\s*\d+\s*$", "", (description or "").strip(), flags=re.IGNORECASE
    ).upper()


async def discover_dvr_channel_names() -> int:
    """Read every DVR's channel names so both cameras of a room are known.

    Half the classrooms were mapped with only their C1 channel, so parents got
    one angle of those rooms while the C2 camera sat unused on the recorder.
    The names are already on the DVR ('G8B C2'), so they are read rather than
    typed in again.
    """
    dvrs = config.get("dvrs", [])

    async def _read_one(dvr_index: int, dvr: dict) -> int:
        ip = dvr.get("ip")
        port = dvr.get("port", 80)
        url = f"http://{ip}:{port}/ISAPI/System/Video/inputs/channels"
        if _credentials_refused(dvr):
            logger.warning(
                "Skipping channel name discovery on %s; it is refusing our "
                "login and another attempt would keep it locked", ip,
            )
            return 0
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    url,
                    auth=httpx.DigestAuth(dvr["username"], dvr["password"]),
                )
            if resp.status_code != 200:
                logger.warning(
                    "Channel name discovery on %s returned HTTP %s",
                    ip, resp.status_code,
                )
                return 0
            names: dict[str, int] = {}
            for block in re.findall(
                r"<VideoInputChannel[^>]*>(.*?)</VideoInputChannel>",
                resp.text,
                flags=re.DOTALL,
            ):
                id_match = re.search(r"<id>\s*(\d+)\s*</id>", block)
                name_match = re.search(r"<name>(.*?)</name>", block, re.DOTALL)
                if not id_match or not name_match:
                    continue
                name = name_match.group(1).strip().upper()
                if name:
                    names[name] = int(id_match.group(1))
            if names:
                _discovered_dvr_channels[dvr_index] = names
            return len(names)
        except Exception as exc:
            logger.warning("Channel name discovery failed for %s: %s", ip, exc)
            return 0

    # The recorders are read in parallel: one unreachable recorder would
    # otherwise hold up the rest for its whole timeout.
    counts = await asyncio.gather(
        *(_read_one(i, dvr) for i, dvr in enumerate(dvrs)),
        return_exceptions=True,
    )
    discovered = sum(c for c in counts if isinstance(c, int))
    logger.info(
        "Discovered %d camera channel name(s) across %d DVR(s)",
        discovered, len(_discovered_dvr_channels),
    )
    return discovered


def _sibling_cameras_from_dvr(
    resolved: list[tuple[dict, int, str]]
) -> list[tuple[dict, int, str]]:
    """Other angles of the same room, taken from the DVR's own channel names."""
    dvrs = config.get("dvrs", [])
    known_channels = {(dvr.get("ip"), channel) for dvr, channel, _ in resolved}
    bases = {_camera_name_base(desc) for _, _, desc in resolved if desc}
    bases.discard("")
    if not bases:
        return []
    extra: list[tuple[dict, int, str]] = []
    for dvr_index, names in _discovered_dvr_channels.items():
        if not (0 <= dvr_index < len(dvrs)):
            continue
        dvr = dvrs[dvr_index]
        for name, channel in names.items():
            if _camera_name_base(name) not in bases:
                continue
            if (dvr.get("ip"), channel) in known_channels:
                continue
            known_channels.add((dvr.get("ip"), channel))
            extra.append((dvr, channel, name))
    return extra


def find_camera_for_classroom(classroom: str) -> tuple[dict, int, str] | None:
    """Look up the DVR, channel number, and description for a given classroom (returns best/first camera)."""
    result = find_all_cameras_for_classroom(classroom)
    if result:
        return result[0]  # Return first (best) camera
    return None


def find_all_cameras_for_classroom(classroom: str) -> list[tuple[dict, int, str]] | None:
    """Look up ALL DVR cameras for a given classroom.

    Returns a list of (dvr_dict, channel, description) tuples for all cameras
    (C1 and C2) mapped to this classroom. Returns None if no cameras found.

    camera_mapping structure:
    {
        "GRADE 3C": {
            "dvr_index": 1, "channel": 17, "description": "G3C C1",
            "all_cameras": [
                {"dvr_index": 1, "channel": 17, "description": "G3C C1", "cam_type": "C1"},
                {"dvr_index": 1, "channel": 13, "description": "G3C C2", "cam_type": "C2"}
            ]
        }
    }
    """
    import re
    mapping = config.get("camera_mapping", {})
    classroom_upper = classroom.strip().upper()
    dvrs = config.get("dvrs", [])

    def _resolve_entry(val: dict) -> list[tuple[dict, int, str]]:
        """Resolve a mapping entry to a list of (dvr, channel, description) tuples."""
        results = []
        all_cams = val.get("all_cameras", [])
        if all_cams:
            for cam in all_cams:
                dvr_idx = cam.get("dvr_index", 0)
                channel = cam.get("channel", 1)
                desc = cam.get("description", "")
                if 0 <= dvr_idx < len(dvrs):
                    results.append((dvrs[dvr_idx], channel, desc))
        else:
            # Single camera entry (no all_cameras field)
            dvr_idx = val.get("dvr_index", 0)
            channel = val.get("channel", 1)
            desc = val.get("description", "")
            if 0 <= dvr_idx < len(dvrs):
                results.append((dvrs[dvr_idx], channel, desc))
        return results or None

    def _strip_dvr_suffix(key: str) -> str:
        """Strip '(DVR X Ch Y)' suffix from key for comparison.
        E.g. 'NUR-3 (DVR 2 Ch 22)' -> 'NUR-3'"""
        return re.sub(r'\s*\(DVR\s*\d+\s*Ch\s*\d+\)\s*$', '', key, flags=re.IGNORECASE).strip().upper()

    def _find_all_vals(target: str) -> list[dict]:
        """Find ALL mapping entries matching the classroom, using fuzzy matching.
        Returns list of mapping dicts for all matching cameras."""
        results = []

        # 0. Match by stripping DVR suffix from keys: "NUR-3 (DVR 2 Ch 22)" -> "NUR-3"
        for key, val in mapping.items():
            key_base = _strip_dvr_suffix(key)
            if key_base == target:
                results.append(val)
        if results:
            return results

        # 1. Direct match (case-insensitive)
        for key, val in mapping.items():
            if key.strip().upper() == target:
                return [val]
        # 2. Whitespace-normalized match (also strip DVR suffix)
        clean = re.sub(r'\s+', '', target)
        for key, val in mapping.items():
            key_base = _strip_dvr_suffix(key)
            key_clean = re.sub(r'\s+', '', key_base)
            if key_clean == clean:
                results.append(val)
        if results:
            return results
        # 3. Strip section letter: "GRADE 6A" → "GRADE 6"
        m = re.match(r'^(GRADE\s*\d{1,2})\s*[A-D]$', target)
        if m:
            grade_no_section_clean = re.sub(r'\s+', '', m.group(1).strip())
            for key, val in mapping.items():
                key_base = _strip_dvr_suffix(key)
                key_clean = re.sub(r'\s+', '', key_base)
                if key_clean == grade_no_section_clean:
                    logger.info(f"Fuzzy camera match: {target} -> {key} (section stripped)")
                    results.append(val)
            if results:
                return results
        # 4. Grade without section: "GRADE 9" → find "GRADE 9A" if it's the only match
        m3 = re.match(r'^GRADE\s*(\d{1,2})$', target)
        if m3:
            grade_num = m3.group(1)
            candidates = []
            for key, val in mapping.items():
                key_base = _strip_dvr_suffix(key)
                km = re.match(r'^GRADE\s*' + grade_num + r'\s*[A-D]$', key_base)
                if km:
                    candidates.append((key, val))
            if len(candidates) == 1:
                logger.info(f"Fuzzy camera match: {target} -> {candidates[0][0]} (section inferred)")
                return [candidates[0][1]]
            elif candidates:
                # Multiple sections exist — pick section A as default
                for key, val in candidates:
                    key_base = _strip_dvr_suffix(key)
                    if key_base.endswith('A'):
                        logger.info(f"Fuzzy camera match: {target} -> {key} (default section A)")
                        return [val]
                logger.info(f"Fuzzy camera match: {target} -> {candidates[0][0]} (first of {len(candidates)})")
                return [candidates[0][1]]
        # 5. Strip number from Nursery/Prep: "NURSERY 4" → "NURSERY"
        m2 = re.match(r'^(NURSERY|NUR|PREP)\s*[-]?\s*\d+$', target)
        if m2:
            base_name = m2.group(1)
            for key, val in mapping.items():
                key_base = _strip_dvr_suffix(key)
                if key_base == base_name:
                    logger.info(f"Fuzzy camera match: {target} -> {key} (number stripped)")
                    return [val]
        return []

    all_vals = _find_all_vals(classroom_upper)
    if all_vals:
        all_results = []
        seen: set[tuple[str, int]] = set()
        for val in all_vals:
            resolved = _resolve_entry(val)
            for dvr, channel, desc in resolved or []:
                key = (dvr.get("ip", ""), channel)
                if key in seen:
                    continue
                seen.add(key)
                all_results.append((dvr, channel, desc))
        if all_results:
            all_results.extend(_sibling_cameras_from_dvr(all_results))
            return all_results

    logger.warning(f"No camera mapping found for: {classroom!r}")
    return None


# ---------------------------------------------------------------------------
# WebSocket client — connects to cloud bot
# ---------------------------------------------------------------------------

ws_connection = None
ws_task = None
_ws_last_activity = 0.0
_ws_disconnected_since = 0.0
_ws_recycles = 0
_WS_STALE_SECONDS = max(
    60.0, float(os.environ.get("WS_STALE_SECONDS", "180"))
)
_snapshot_tasks: set[asyncio.Task] = set()
_live_requests_in_flight = 0
# One family's classroom must not sit in a queue behind another family's:
# each recorder already has its own concurrency limiter, so several parent
# requests can be served at once without hammering any one DVR.
_snapshot_request_semaphore = asyncio.Semaphore(
    max(1, int(os.environ.get("SNAPSHOT_CONCURRENT_REQUESTS", "6")))
)
_SNAPSHOT_CAMERA_TIMEOUT_SECONDS = max(
    1.0, float(os.environ.get("SNAPSHOT_CAMERA_TIMEOUT_SECONDS", "40"))
)


def _snapshot_task_done(task: asyncio.Task) -> None:
    _snapshot_tasks.discard(task)
    try:
        task.result()
    except Exception as exc:
        logger.error(
            "Snapshot request task failed: %s",
            _exception_text(exc),
            exc_info=True,
        )


def _note_ws_activity() -> None:
    global _ws_last_activity, _ws_disconnected_since
    _ws_last_activity = time.monotonic()
    _ws_disconnected_since = 0.0


def _ws_looks_connected() -> bool:
    ws = ws_connection
    return ws is not None and getattr(ws, "open", False)


def ws_link_health() -> dict:
    """What the agent believes about its link to the cloud."""
    silent_for = (
        round(time.monotonic() - _ws_last_activity, 1)
        if _ws_last_activity
        else None
    )
    return {
        "connected": _ws_looks_connected(),
        "silent_seconds": silent_for,
        "recycles": _ws_recycles,
    }


async def _recycle_websocket(reason: str) -> None:
    """Tear the cloud link down and build a fresh one.

    The socket can look open to us long after the cloud has stopped seeing
    us, and parents' requests then land nowhere. Rebuilding is cheap.
    """
    global ws_connection, ws_task, _ws_recycles
    _ws_recycles += 1
    logger.error("WATCHDOG: recycling the cloud link (%s)", reason)
    task = ws_task
    ws_connection = None
    ws_task = None
    if task is not None and not task.done():
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass
        except Exception as exc:
            logger.debug("Old cloud link ended with %s", _exception_text(exc))
    ws_task = asyncio.create_task(websocket_client())
    _note_ws_activity()


async def _cloud_says_we_are_connected() -> bool | None:
    """The cloud's own view of our link, None when we cannot ask."""
    api_url = (
        attendance_engine.whatsapp_api_url
        or "https://ppis-whatsapp-bot.fly.dev"
    )
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            response = await client.get(f"{api_url}/api/agent/health")
            if response.status_code != 200:
                return None
            return bool(response.json().get("connected"))
    except Exception:
        return None


async def _repair_cloud_link_if_needed() -> None:
    """Force a reconnect when the cloud cannot see us any more."""
    global _ws_disconnected_since
    if not _ws_looks_connected():
        if not _ws_disconnected_since:
            _ws_disconnected_since = time.monotonic()
        elif time.monotonic() - _ws_disconnected_since > _WS_STALE_SECONDS:
            await _recycle_websocket("no cloud link for over 3 minutes")
        return
    if await _cloud_says_we_are_connected() is False:
        # Our socket looks fine but the cloud has stopped seeing us, so the
        # connection is half-open and every parent request is being lost.
        await _recycle_websocket("cloud reports the agent as offline")


async def websocket_client():
    """Persistent WebSocket connection to the cloud bot.
    Receives snapshot requests and sends back images."""
    global ws_connection
    url = config.get("cloud_bot_url", "wss://ppis-whatsapp-bot.fly.dev/ws/agent")
    secret = config.get("agent_secret", os.environ.get("AGENT_SECRET", ""))

    ws_backoff = 5
    while True:
        try:
            logger.info(f"Connecting to cloud bot WebSocket: {url}")
            async with websockets.connect(
                url,
                extra_headers={"X-Agent-Secret": secret},
                ping_interval=30,
                ping_timeout=30,
                max_size=10 * 1024 * 1024,  # 10 MB max message size
            ) as ws:
                ws_connection = ws
                ws_backoff = 5  # Reset backoff on successful connect
                _note_ws_activity()
                logger.info("Connected to cloud bot WebSocket")

                # Send hello
                await ws.send(json.dumps({
                    "type": "agent_hello",
                    "dvr_count": len(config.get("dvrs", [])),
                    "camera_count": len(config.get("camera_mapping", {})),
                    "dvr_health": dvr_snapshot_health(),
                    "code_commit": _running_commit(),
                    "started_at_ist": _process_started_at_ist(),
                }))

                async for message in ws:
                    _note_ws_activity()
                    try:
                        data = json.loads(message)
                        msg_type = data.get("type", "")

                        if msg_type == "snapshot_request":
                            classroom = data.get("classroom", "")
                            request_id = data.get("request_id", "")
                            logger.info(f"Snapshot request for classroom: {classroom} (req: {request_id})")
                            task = asyncio.create_task(
                                handle_snapshot_request(ws, classroom, request_id)
                            )
                            _snapshot_tasks.add(task)
                            task.add_done_callback(_snapshot_task_done)

                        elif msg_type == "ping":
                            await ws.send(json.dumps({
                                "type": "pong",
                                "dvr_health": dvr_snapshot_health(),
                            }))

                        elif msg_type == "test_connection":
                            dvr_idx = data.get("dvr_index", 0)
                            dvrs = config.get("dvrs", [])
                            if 0 <= dvr_idx < len(dvrs):
                                result = await test_dvr_connection(dvrs[dvr_idx])
                                await ws.send(json.dumps({"type": "test_result", **result}))

                        elif msg_type == "test_all_dvrs":
                            request_id = data.get("request_id", "")
                            dvrs = config.get("dvrs", [])
                            results = []
                            for i, dvr in enumerate(dvrs):
                                result = await test_dvr_connection(dvr)
                                result["dvr_index"] = i
                                result["name"] = dvr.get("name", f"DVR {i}")
                                results.append(result)
                            await ws.send(json.dumps({
                                "type": "test_all_dvrs_result",
                                "request_id": request_id,
                                "results": results,
                            }))

                        elif msg_type == "update_camera_mapping":
                            new_mapping = data.get("camera_mapping", {})
                            if new_mapping:
                                config["camera_mapping"] = new_mapping
                                save_config(config)
                                logger.info(f"Camera mapping updated remotely: {len(new_mapping)} entries")
                                await ws.send(json.dumps({
                                    "type": "mapping_updated",
                                    "success": True,
                                    "count": len(new_mapping),
                                }))
                            else:
                                await ws.send(json.dumps({
                                    "type": "mapping_updated",
                                    "success": False,
                                    "error": "Empty mapping data",
                                }))

                        elif msg_type == "update_dvrs":
                            new_dvrs = data.get("dvrs", [])
                            if new_dvrs:
                                config["dvrs"] = new_dvrs
                                save_config(config)
                                logger.info(f"DVRs updated remotely: {len(new_dvrs)} entries")
                                await ws.send(json.dumps({
                                    "type": "dvrs_updated",
                                    "success": True,
                                    "count": len(new_dvrs),
                                }))
                            else:
                                await ws.send(json.dumps({
                                    "type": "dvrs_updated",
                                    "success": False,
                                    "error": "Empty DVR data",
                                }))

                        elif msg_type == "restart":
                            logger.info("REMOTE RESTART requested via WebSocket")
                            await ws.send(json.dumps({
                                "type": "restart_ack",
                                "success": True,
                            }))
                            await ws.close()
                            os._exit(0)  # run_forever.bat will restart with git pull

                        elif msg_type == "sync_faces":
                            logger.info("REMOTE FACE SYNC requested via WebSocket")
                            synced = await sync_faces_from_cloud()
                            if synced > 0:
                                await _reload_faces_off_loop()
                            await ws.send(json.dumps({
                                "type": "sync_faces_result",
                                "synced": synced,
                            }))

                        else:
                            logger.warning(f"Unknown WS message type: {msg_type}")

                    except json.JSONDecodeError:
                        logger.error(f"Invalid JSON from cloud bot: {message[:200]}")
                    except Exception as e:
                        logger.error(
                            "Error handling WS message: %s",
                            _exception_text(e),
                            exc_info=True,
                        )

        except websockets.exceptions.ConnectionClosed as e:
            logger.warning(
                "WebSocket closed: %s. Reconnecting in 5s...",
                _exception_text(e),
            )
            ws_connection = None
            ws_backoff = 5
        except Exception as e:
            logger.error(
                "WebSocket error: %s. Reconnecting in %ss...",
                _exception_text(e),
                ws_backoff,
            )
            ws_connection = None

        await asyncio.sleep(ws_backoff)
        ws_backoff = min(ws_backoff * 2, 60)


_IST = timezone(timedelta(hours=5, minutes=30))
_DAYLIGHT_HOURS_IST = (7, 18)
_COLOUR_REPAIR_ATTEMPTED: dict[tuple[str, int], str] = {}


def _image_has_no_colour(data: bytes) -> bool | None:
    """True when a JPEG carries no colour, i.e. the camera is in night mode.

    Returns None when the picture cannot be inspected.
    """
    if Image is None:
        return None
    try:
        img = Image.open(io.BytesIO(data))
        if img.mode in ("L", "1", "I", "F"):
            return True
        pixels = list(img.convert("RGB").resize((64, 36)).getdata())
    except Exception as exc:
        logger.debug("Colour check failed: %s", exc)
        return None
    if not pixels:
        return None
    coloured = sum(
        1 for r, g, b in pixels if max(r, g, b) - min(r, g, b) > 24
    )
    return coloured <= len(pixels) * 0.02


async def _restore_daylight_colour(
    dvr: dict, channel: int, desc: str
) -> tuple[bool, str]:
    """Take a camera out of night mode so daytime pictures are in colour.

    Returns (setting accepted, day/night mode the recorder reported).
    """
    ip = dvr["ip"]
    url = (
        f"http://{ip}:{dvr.get('port', 80)}"
        f"/ISAPI/Image/channels/{channel}/ircutFilter"
    )
    if _credentials_refused(dvr):
        return False, "login refused"
    auth = httpx.DigestAuth(dvr["username"], dvr["password"])
    body = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<IrcutFilter><IrcutFilterType>auto</IrcutFilterType>"
        "<nightToDayFilterLevel>4</nightToDayFilterLevel>"
        "</IrcutFilter>"
    )
    reported = "unknown"
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            current = await client.get(url, auth=auth)
            if current.status_code == 200:
                mode = re.search(
                    r"<IrcutFilterType>(.*?)</IrcutFilterType>", current.text
                )
                if mode:
                    reported = mode.group(1).strip().lower()
                logger.warning(
                    "%s (ch%d on %s) returned a colourless daytime picture; "
                    "recorder reports day/night mode %s",
                    desc, channel, ip, reported,
                )
            else:
                logger.warning(
                    "%s (ch%d on %s) returned a colourless daytime picture; "
                    "day/night settings unreadable (HTTP %s)",
                    desc, channel, ip, current.status_code,
                )
            resp = await client.put(
                url,
                content=body,
                headers={"Content-Type": "application/xml"},
                auth=auth,
            )
    except Exception as exc:
        logger.warning(
            "Could not restore daylight colour on %s ch%d: %s", ip, channel, exc
        )
        return False, reported
    if resp.status_code == 200:
        logger.info(
            "Set %s (ch%d on %s) back to automatic day/night", desc, channel, ip
        )
        return True, reported
    logger.warning(
        "Recorder refused the day/night change for %s ch%d (HTTP %s): %s",
        desc, channel, resp.status_code, resp.text[:200],
    )
    return False, reported


async def _repair_colour_if_night_mode(
    snapshot: bytes,
    camera: tuple[dict, int, str],
) -> bytes:
    """Recapture in colour when a camera is stuck in night mode during the day.

    Only ever tried once per camera per day, and only in daylight hours (IST).
    """
    dvr, channel, desc = camera
    now = datetime.now(_IST)
    if not _DAYLIGHT_HOURS_IST[0] <= now.hour < _DAYLIGHT_HOURS_IST[1]:
        return snapshot
    key = (dvr["ip"], channel)
    today = now.strftime("%Y-%m-%d")
    if _COLOUR_REPAIR_ATTEMPTED.get(key) == today:
        return snapshot
    if _image_has_no_colour(snapshot) is not True:
        return snapshot

    _COLOUR_REPAIR_ATTEMPTED[key] = today
    deadline = _live_request_deadline.get()
    if deadline is not None and deadline - time.monotonic() < 20.0:
        # Not enough of the parent's request budget left to wait for a colour
        # recapture: fix the camera in the background so the next photo is fine.
        asyncio.create_task(_restore_daylight_colour(dvr, channel, desc))
        return snapshot
    changed, reported_mode = await _restore_daylight_colour(dvr, channel, desc)
    if not changed:
        return snapshot
    await asyncio.sleep(2.0)
    try:
        retry = await asyncio.wait_for(capture_snapshot(dvr, channel), timeout=15.0)
    except asyncio.TimeoutError:
        return snapshot
    if retry and _image_has_no_colour(retry) is not True:
        logger.info("Recaptured %s in colour after leaving night mode", desc)
        return retry
    if reported_mode == "auto":
        # The camera was already deciding for itself, so it chose infrared
        # because the room was too dim — lights, not settings.
        logger.warning(
            "%s is on automatic day/night and still shooting infrared: the room "
            "light level is too low for a colour picture",
            desc,
        )
    else:
        logger.warning(
            "%s is still colourless after the day/night reset (recorder mode "
            "%s) — its infrared filter likely needs attention on the camera "
            "itself",
            desc, reported_mode,
        )
    return retry or snapshot


async def _capture_classroom_camera(
    classroom: str,
    camera: tuple[dict, int, str],
) -> tuple[bytes, str, str, dict] | None:
    dvr, channel, desc = camera
    report: dict = {}
    _live_capture_report.set(report)
    try:
        request_deadline = _live_request_deadline.get()
        timeout = _SNAPSHOT_CAMERA_TIMEOUT_SECONDS
        if request_deadline is not None:
            timeout = max(0.1, min(timeout, request_deadline - time.monotonic()))
        snapshot = await asyncio.wait_for(
            capture_snapshot(dvr, channel),
            timeout=timeout,
        )
    except asyncio.TimeoutError:
        logger.warning(
            "Snapshot timed out after %.1fs from DVR %s channel %d (%s)",
            timeout,
            dvr["ip"],
            channel,
            desc,
        )
        return None
    if not snapshot:
        logger.warning(
            f"Failed to capture from DVR {dvr['ip']} channel {channel} ({desc})"
        )
        return None

    snapshot = await _repair_colour_if_night_mode(snapshot, camera)

    ts = int(time.time() * 1000)
    cam_label = desc.split()[-1] if desc else f"ch{channel}"
    safe_name = classroom.replace(' ', '_').replace('/', '_').replace('\\', '_')
    filename = f"{safe_name}_{cam_label}_{ts}.jpg"
    filepath = SNAPSHOT_DIR / filename
    with open(filepath, "wb") as file:
        file.write(snapshot)
    logger.info(f"Snapshot captured: {filename} ({len(snapshot)} bytes) - {desc}")
    return snapshot, filename, desc, report


async def handle_snapshot_request(ws, classroom: str, request_id: str):
    """Handle a snapshot request from the cloud bot.

    Captures from ALL cameras (C1 and C2) for the classroom.
    Sends each image as a separate WebSocket message to avoid size limits.
    Protocol:
      1. snapshot_image  (one per captured image, sent individually)
      2. snapshot_complete (final message with total count)
    Falls back to legacy single-message format if only 1 image captured.
    """
    global _live_requests_in_flight
    queued_at = time.monotonic()
    _live_requests_in_flight += 1
    try:
        await _serve_snapshot_request(ws, classroom, request_id, queued_at)
    finally:
        _live_requests_in_flight -= 1


async def _serve_snapshot_request(
    ws, classroom: str, request_id: str, queued_at: float
):
    async with _snapshot_request_semaphore:
        waited = time.monotonic() - queued_at
        if waited >= 1.0:
            logger.warning(
                "Snapshot for %s waited %.1fs for a free capture slot",
                classroom,
                waited,
            )
        await _handle_snapshot_request(ws, classroom, request_id)


async def _handle_snapshot_request(ws, classroom: str, request_id: str):
    request_started = time.monotonic()
    request_token = _live_request_deadline.set(
        request_started + _SNAPSHOT_LIVE_REQUEST_BUDGET_SECONDS
    )
    classroom_token = _live_request_classroom.set(classroom)
    all_cameras = find_all_cameras_for_classroom(classroom)

    if not all_cameras:
        await ws.send(json.dumps({
            "type": "snapshot_response",
            "request_id": request_id,
            "success": False,
            "error": f"No camera mapped for classroom: {classroom}",
        }))
        _live_request_deadline.reset(request_token)
        _live_request_classroom.reset(classroom_token)
        return

    logger.info(f"Capturing from {len(all_cameras)} camera(s) for {classroom}")

    capture_tasks = [
        asyncio.create_task(_capture_classroom_camera(classroom, camera))
        for camera in all_cameras[:2]
    ]
    expected_total = len(capture_tasks)
    sent_count = 0

    for completed in asyncio.as_completed(capture_tasks):
        result = await completed
        if result is None:
            continue

        raw_data, filename, desc, capture = result
        compressed = compress_jpeg(
            raw_data,
            max_bytes=_LIVE_SNAPSHOT_MAX_BYTES,
            quality_start=_LIVE_SNAPSHOT_JPEG_QUALITY,
        )
        b64 = base64.b64encode(compressed).decode("ascii")
        width, height = _jpeg_size(compressed)
        await ws.send(json.dumps({
            "type": "snapshot_image",
            "request_id": request_id,
            "classroom": classroom,
            "image_index": sent_count,
            "image_total": expected_total,
            "filename": filename,
            "image_base64": b64,
            "size_bytes": len(compressed),
            "width": width,
            "height": height,
            "description": desc,
            "capture": capture,
        }))
        sent_count += 1
        logger.info(
            "Sent image %d/%d: %s (%d bytes, %dx%d) - %s in %.2fs",
            sent_count,
            expected_total,
            filename,
            len(compressed),
            width,
            height,
            desc,
            time.monotonic() - request_started,
        )

    if sent_count == 0:
        await ws.send(json.dumps({
            "type": "snapshot_response",
            "request_id": request_id,
            "success": False,
            "error": f"Failed to capture snapshot from any camera for {classroom}",
            "detail": _last_capture_error,
        }))
        _live_request_deadline.reset(request_token)
        _live_request_classroom.reset(classroom_token)
        return

    # Send completion message
    await ws.send(json.dumps({
        "type": "snapshot_complete",
        "request_id": request_id,
        "success": True,
        "classroom": classroom,
        "image_count": sent_count,
    }))
    logger.info(
        "Sent snapshot_complete for %s: %d image(s) in %.2fs",
        classroom,
        sent_count,
        time.monotonic() - request_started,
    )
    _live_request_deadline.reset(request_token)
    _live_request_classroom.reset(classroom_token)


# ---------------------------------------------------------------------------
# FastAPI — Local web UI
# ---------------------------------------------------------------------------

async def _auto_start_classwise():
    """Auto-start classwise monitoring after a brief delay.

    This ensures the system is always-on without manual intervention.
    """
    try:
        await asyncio.sleep(10)  # Let other startup tasks finish
        if attendance_engine.classwise_running or attendance_engine.running:
            logger.info("Monitoring already active — skipping auto-start")
            return

        dvrs = config.get("dvrs", [])
        camera_mapping = config.get("camera_mapping", {})
        if not dvrs or not camera_mapping:
            logger.warning("Auto-start skipped: no DVRs or camera mapping configured")
            return

        attendance_engine.test_mode = False

        # Configure camera alert phones from agent settings
        import database as db_mod
        alert_phones_str = db_mod.get_attendance_setting("camera_alert_phones", "")
        if alert_phones_str:
            attendance_engine._admin_phones = [p.strip() for p in alert_phones_str.split(",") if p.strip()]
            logger.info(f"Camera alerts configured for: {attendance_engine._admin_phones}")

        attendance_engine.classwise_running = True
        attendance_engine._classwise_task = asyncio.create_task(
            attendance_engine.classwise_monitoring_loop(dvrs, camera_mapping)
        )
        logger.info("AUTO-START: Classwise attendance monitoring started automatically")
    except Exception as e:
        logger.error(f"AUTO-START FAILED: {e}", exc_info=True)


async def _auto_start_mood_and_sighting():
    """Auto-start mood detection and teacher sighting tracker after delay."""
    try:
        await asyncio.sleep(15)  # Let face sync and classwise start first

        dvrs = config.get("dvrs", [])
        camera_mapping = config.get("camera_mapping", {})
        if not dvrs or not camera_mapping:
            logger.warning("Mood/Sighting auto-start skipped: no DVRs or camera mapping")
            return

        agent_secret = config.get("agent_secret", os.environ.get("AGENT_SECRET", ""))
        mood_detector.agent_secret = agent_secret
        sighting_tracker.agent_secret = agent_secret

        mood_detector.start(dvrs, camera_mapping)
        sighting_tracker.start(dvrs, camera_mapping)
        logger.info("AUTO-START: Mood detection + teacher sighting tracker started")
    except Exception as e:
        logger.error(f"Mood/Sighting auto-start failed: {e}", exc_info=True)


def _restart_websocket_task_if_needed() -> bool:
    global ws_connection, ws_task
    if ws_task is not None and not ws_task.done():
        return False

    if ws_task is not None:
        try:
            error = ws_task.exception()
        except (asyncio.CancelledError, asyncio.InvalidStateError):
            error = None
        logger.error("WATCHDOG: WebSocket task stopped — restarting (%s)", error)

    ws_connection = None
    ws_task = asyncio.create_task(websocket_client())
    return True


async def _health_watchdog():
    """Background watchdog that monitors system health and auto-recovers.

    Checks every 60 seconds:
    - Is classwise monitoring running? If not, restart it.
    - Are cameras responding?
    - Is the notification system reachable?
    - Periodically sync new faces from cloud.
    - Monitor memory usage and force cleanup if too high.
    - Clean stale snapshot files.
    """
    import gc
    face_sync_counter = 0
    cleanup_counter = 0
    while True:
        await asyncio.sleep(60)
        try:
            if not _restart_websocket_task_if_needed():
                await _repair_cloud_link_if_needed()

            # --- Check 1: Classwise monitoring alive ---
            if attendance_engine._health.get("auto_start_enabled", True):
                if not attendance_engine.classwise_running and not attendance_engine.running:
                    dvrs = config.get("dvrs", [])
                    camera_mapping = config.get("camera_mapping", {})
                    if dvrs and camera_mapping:
                        logger.warning("WATCHDOG: Classwise monitoring stopped — restarting")
                        attendance_engine._health["total_recoveries"] += 1
                        attendance_engine.test_mode = False
                        await _reload_faces_off_loop()
                        attendance_engine.classwise_running = True
                        attendance_engine._classwise_task = asyncio.create_task(
                            attendance_engine.classwise_monitoring_loop(dvrs, camera_mapping)
                        )

            # --- Check 2: Notification system reachable ---
            try:
                api_url = attendance_engine.whatsapp_api_url or "https://ppis-whatsapp-bot.fly.dev"
                async with httpx.AsyncClient(timeout=10) as client:
                    resp = await client.get(f"{api_url}/debug/version")
                    if resp.status_code == 200:
                        attendance_engine._health["notification_system"] = "ok"
                    else:
                        attendance_engine._health["notification_system"] = "degraded"
            except Exception:
                attendance_engine._health["notification_system"] = "error"

            # --- Check 3: Periodic face sync from cloud (every 5 min) ---
            face_sync_counter += 1
            if face_sync_counter >= 5:
                face_sync_counter = 0
                synced = await sync_faces_from_cloud()
                if synced > 0:
                    await _reload_faces_off_loop()
                    logger.info(f"WATCHDOG: Synced {synced} new face(s) from cloud")

            # --- Check 4: Tiered memory management (every 2 min) ---
            cleanup_counter += 1
            if cleanup_counter >= 2:
                cleanup_counter = 0
                try:
                    import psutil
                    proc = psutil.Process()
                    mem_mb = proc.memory_info().rss / (1024 * 1024)
                    attendance_engine._health["memory_mb"] = round(mem_mb, 1)

                    # Tier 1: Normal cleanup (>400MB)
                    if mem_mb > 400:
                        stats = attendance_engine.cleanup_memory(aggressive=False)
                        logger.info(f"WATCHDOG: Memory {mem_mb:.0f}MB — routine cleanup: {stats}")
                        _cleanup_old_snapshots()

                    # Tier 2: Aggressive cleanup (>700MB)
                    if mem_mb > 700:
                        stats = attendance_engine.cleanup_memory(aggressive=True)
                        logger.warning(f"WATCHDOG: High memory {mem_mb:.0f}MB — aggressive cleanup: {stats}")

                    # Tier 3: Critical — force restart (>1200MB)
                    if mem_mb > 1200:
                        logger.critical(
                            f"WATCHDOG: CRITICAL memory {mem_mb:.0f}MB — "
                            f"forcing process restart to prevent OOM"
                        )
                        attendance_engine.add_debug_log(
                            "memory_restart",
                            f"Process restarting due to critical memory: {mem_mb:.0f}MB"
                        )
                        # Exit with code 1 — run_forever.bat will auto-restart
                        import sys
                        sys.exit(1)

                except ImportError:
                    pass
                except SystemExit:
                    raise
                except Exception as e:
                    logger.debug(f"WATCHDOG: Memory check error: {e}")

            # --- Check 5: Clean stale snapshots (every 30 min) ---
            if face_sync_counter == 0 and cleanup_counter == 0:
                _cleanup_old_snapshots()

            attendance_engine._health["last_health_check"] = (
                __import__("datetime").datetime.now().isoformat()
            )
        except Exception as e:
            logger.error(f"WATCHDOG error: {e}")


def _cleanup_old_snapshots():
    """Remove snapshot files older than 2 hours to prevent disk fill."""
    cutoff = time.time() - 7200  # 2 hours
    for snap_dir in [SNAPSHOT_DIR, Path(__file__).parent / "attendance_snapshots"]:
        if not snap_dir.exists():
            continue
        try:
            for f in snap_dir.iterdir():
                if f.is_file() and f.stat().st_mtime < cutoff:
                    f.unlink()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Periodic Entry Gate Snapshot → WhatsApp (every 10 minutes)
# ---------------------------------------------------------------------------
_GATE_SNAPSHOT_INTERVAL = 600  # 10 minutes
_GATE_SNAPSHOT_CAMERA_IP = "192.168.0.14"  # DVR 3
_GATE_SNAPSHOT_CHANNEL = 20               # Entry Gate-1
_GATE_SNAPSHOT_API = f"{CLOUD_API_BASE}/api/gate/entry-gate-snapshot"


async def _entry_gate_snapshot_loop():
    """Capture Entry Gate-1 snapshot every 10 minutes and send to WhatsApp via backend."""
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    IST = _tz(_td(hours=5, minutes=30))

    await asyncio.sleep(30)  # initial delay to let config load
    logger.info("[GATE-SNAP] Entry gate snapshot loop started (every %ds)", _GATE_SNAPSHOT_INTERVAL)

    while True:
        try:
            now = _dt.now(IST)
            hour = now.hour
            # Only send snapshots during school hours (7 AM - 5 PM IST)
            if hour < 7 or hour >= 17:
                logger.debug("[GATE-SNAP] Outside school hours (%d), skipping", hour)
                await asyncio.sleep(_GATE_SNAPSHOT_INTERVAL)
                continue

            # Find DVR 3 in config
            dvr = None
            for d in config.get("dvrs", []):
                if d.get("ip") == _GATE_SNAPSHOT_CAMERA_IP:
                    dvr = d
                    break

            if not dvr:
                logger.warning("[GATE-SNAP] DVR 3 (%s) not found in config", _GATE_SNAPSHOT_CAMERA_IP)
                await asyncio.sleep(_GATE_SNAPSHOT_INTERVAL)
                continue

            frame = await capture_snapshot(dvr, _GATE_SNAPSHOT_CHANNEL)
            if not frame:
                logger.warning("[GATE-SNAP] Failed to capture Entry Gate-1 snapshot")
                await asyncio.sleep(_GATE_SNAPSHOT_INTERVAL)
                continue

            # Compress if needed
            frame = compress_jpeg(frame, max_bytes=200_000)

            image_b64 = base64.b64encode(frame).decode("ascii")
            ts = now.strftime("%d-%m-%Y %H:%M:%S IST")
            payload = {
                "image_b64": image_b64,
                "camera": "ENTRY GATE-1",
                "timestamp": ts,
            }

            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.post(_GATE_SNAPSHOT_API, json=payload)
                if resp.status_code == 200:
                    logger.info("[GATE-SNAP] Entry Gate snapshot sent to WhatsApp at %s", ts)
                else:
                    logger.warning("[GATE-SNAP] Backend returned %d: %s", resp.status_code, resp.text[:200])

        except Exception as e:
            logger.error("[GATE-SNAP] Error in snapshot loop: %s", e)

        await asyncio.sleep(_GATE_SNAPSHOT_INTERVAL)


@asynccontextmanager
async def lifespan(app: FastAPI):
    global ws_task, config

    from process_priority import set_windows_process_priority

    set_windows_process_priority("BELOW_NORMAL_PRIORITY_CLASS", "Campus agent")

    # Set up asyncio exception handler to log unhandled task errors
    def _handle_task_exception(loop, context):
        exc = context.get("exception")
        msg = context.get("message", "")
        logger.critical(f"ASYNCIO UNHANDLED: {msg} — {exc}", exc_info=exc)
    asyncio.get_event_loop().set_exception_handler(_handle_task_exception)

    # Initialize database (creates tables including attendance tables)
    import database as db_mod
    db_mod.init_db()
    # Load config from cloud DB (falls back to local config.json)
    config = await load_config()
    # ALWAYS enforce the correct cloud bot URL (prevents stale config.json issues)
    config["cloud_bot_url"] = "wss://ppis-whatsapp-bot.fly.dev/ws/agent"
    logger.info(
        f"Config loaded: {len(config.get('dvrs', []))} DVRs, "
        f"{len(config.get('camera_mapping', {}))} camera mappings"
    )
    # Start WebSocket client FIRST — connects to backend immediately
    # so snapshot requests work even during face sync
    logger.info("Starting WebSocket client task...")
    _restart_websocket_task_if_needed()

    # Read the DVRs' channel names in the background: an early snapshot request
    # is served from the mapping and only gains the second angle once this ends.
    async def _discover_channels_in_background():
        try:
            await discover_dvr_channel_names()
        except Exception as e:
            logger.error(f"Channel name discovery failed (non-fatal): {e}")

    asyncio.create_task(_discover_channels_in_background())

    # Clean up known junk/duplicate face entries before syncing
    cleanup_junk_face_entries()

    # Sync face registrations from cloud DB (downloads images, computes encodings)
    logger.info("Starting face sync from cloud...")
    try:
        await sync_faces_from_cloud()
        logger.info("Face sync completed successfully")
    except Exception as e:
        logger.error(f"Face sync crashed during startup (non-fatal): {e}", exc_info=True)
    # Pre-load registered faces into attendance engine
    logger.info("Loading face encodings into attendance engine...")
    try:
        await _reload_faces_off_loop()
        logger.info("Face encodings loaded")
    except Exception as e:
        logger.error(f"Face reload crashed during startup (non-fatal): {e}", exc_info=True)

    # Auto-start classwise monitoring after brief delay (24/7 always-on)
    asyncio.create_task(_auto_start_classwise())
    # Auto-start mood detection and teacher sighting after delay
    asyncio.create_task(_auto_start_mood_and_sighting())
    # Start health watchdog after 60 seconds (auto-recovery, face sync)
    async def _delayed_watchdog():
        try:
            await asyncio.sleep(60)
            await _health_watchdog()
        except Exception as e:
            logger.error(f"Health watchdog failed: {e}", exc_info=True)

    asyncio.create_task(_delayed_watchdog())
    # Re-try a locked-out recorder by itself once its lock has had silence to
    # expire, so nobody has to reboot a DVR to get its classrooms back.
    asyncio.create_task(_unlock_watch_loop())
    # Pick up merged fixes without anyone running a restart script.
    asyncio.create_task(_auto_update_loop())
    # Periodic Entry Gate snapshots DISABLED per user request (2026-05-30).
    # User only wants unknown person alerts, not routine snapshots.
    # asyncio.create_task(_entry_gate_snapshot_loop())
    logger.info("PPIS Campus Agent started (24/7 mode with auto-recovery)")
    try:
        yield
    except Exception as e:
        logger.critical(f"LIFESPAN CRASH: {e}", exc_info=True)
    finally:
        # Shutdown
        attendance_engine.stop()
        mood_detector.stop()
        sighting_tracker.stop()
        if ws_task:
            ws_task.cancel()
    logger.info("PPIS Campus Agent stopped")


app = FastAPI(title="PPIS Campus Agent", lifespan=lifespan)

# --- TrueFace 3000 ADMS integration ---
from trueface_adms import router as trueface_router
app.include_router(trueface_router)

# Ensure static directories exist
(Path(__file__).parent / "static").mkdir(exist_ok=True)
(Path(__file__).parent / "face_images").mkdir(exist_ok=True)
(Path(__file__).parent / "attendance_snapshots").mkdir(exist_ok=True)

app.mount("/static", StaticFiles(directory=str(Path(__file__).parent / "static")), name="static")
app.mount("/snapshots", StaticFiles(directory=str(SNAPSHOT_DIR)), name="snapshots")
app.mount("/face_images", StaticFiles(directory=str(Path(__file__).parent / "face_images")), name="face_images")
app.mount("/attendance_snapshots", StaticFiles(directory=str(Path(__file__).parent / "attendance_snapshots")), name="attendance_snapshots")


@app.get("/", response_class=HTMLResponse)
async def dashboard():
    """Main dashboard page."""
    return get_dashboard_html()


@app.get("/api/config")
async def get_config():
    """Return current config (without passwords)."""
    safe_dvrs = []
    for d in config.get("dvrs", []):
        safe_dvrs.append({
            "name": d.get("name", ""),
            "ip": d.get("ip", ""),
            "port": d.get("port", 80),
            "username": d.get("username", ""),
            "channels": d.get("channels", 64),
        })
    return {
        "dvrs": safe_dvrs,
        "camera_mapping": config.get("camera_mapping", {}),
        "cloud_bot_url": config.get("cloud_bot_url", ""),
        "config_source": "cloud" if config.get("_from_cloud") else "local",
        "ws_connected": ws_connection is not None and ws_connection.open if ws_connection else False,
        "ws_link": ws_link_health(),
    }


@app.post("/api/dvr/save")
async def save_dvr_config(request: Request):
    """Save DVR configuration locally and sync to cloud.

    Preserves existing passwords when incoming DVR entries have empty passwords
    (the frontend strips passwords for display security).
    """
    body = await request.json()
    dvrs = body.get("dvrs", [])

    # Merge: preserve stored passwords when incoming password is empty
    # Match by (ip, port) key instead of index to handle DVR reordering/deletion
    existing_dvrs = config.get("dvrs", [])
    existing_pw_map = {}
    for d in existing_dvrs:
        key = (d.get("ip", ""), d.get("port", 80))
        existing_pw_map[key] = d.get("password", "")
    for new_dvr in dvrs:
        if not new_dvr.get("password"):
            key = (new_dvr.get("ip", ""), new_dvr.get("port", 80))
            if key in existing_pw_map:
                new_dvr["password"] = existing_pw_map[key]

    config["dvrs"] = dvrs
    save_config(config)
    # Sync to cloud DB
    cloud_synced = False
    try:
        agent_secret = os.environ.get("AGENT_SECRET", "")
        headers = {"X-Agent-Secret": agent_secret} if agent_secret else {}
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                f"{CLOUD_API_BASE}/api/agent-config/dvrs",
                json={"dvrs": dvrs},
                headers=headers,
            )
            cloud_synced = resp.status_code == 200
    except Exception as e:
        logger.warning(f"Failed to sync DVRs to cloud: {e}")
    return {"status": "ok", "dvr_count": len(dvrs), "cloud_synced": cloud_synced}


@app.post("/api/dvr/test/{dvr_index}")
async def test_dvr(dvr_index: int):
    """Test connection to a specific DVR."""
    dvrs = config.get("dvrs", [])
    if dvr_index < 0 or dvr_index >= len(dvrs):
        return JSONResponse({"status": "error", "error": "Invalid DVR index"}, status_code=400)
    result = await test_dvr_connection(dvrs[dvr_index])
    return result


@app.post("/api/snapshot/{classroom}")
async def take_snapshot(classroom: str):
    """Manually capture a snapshot for a classroom."""
    result = find_camera_for_classroom(classroom)
    if not result:
        return JSONResponse(
            {"status": "error", "error": f"No camera mapped for: {classroom}"},
            status_code=404,
        )

    dvr, channel, _desc = result
    snapshot = await capture_snapshot(dvr, channel)
    if not snapshot:
        return JSONResponse(
            {"status": "error", "error": "Failed to capture snapshot"},
            status_code=500,
        )

    ts = int(time.time())
    safe_name = classroom.replace(' ', '_').replace('/', '_').replace('\\', '_')
    filename = f"{safe_name}_{ts}.jpg"
    filepath = SNAPSHOT_DIR / filename
    with open(filepath, "wb") as f:
        f.write(snapshot)

    return {"status": "ok", "filename": filename, "size_bytes": len(snapshot)}


def _parse_camera_xls(file_path: str | Path) -> dict:
    """Parse the PPIS camera Excel (.xls) with side-by-side NVR layout.

    The Excel has this structure:
      Cols A-D: NVR 1 data (S.NO, CAMERA NAME, CAMERA LOCATION, SOUND)
      Col E: empty separator
      Cols F-I: NVR 2 data (S.NO, CAMERA NAME, CAMERA LOCATION, SOUND)
    NVR 3 starts further down in cols A-D after NVR 1 ends.

    Returns a dict mapping camera_name -> {dvr_index, channel, location, sound}.
    """
    import re
    import xlrd

    wb = xlrd.open_workbook(str(file_path))
    ws = wb.sheet_by_index(0)

    mapping = {}
    current_nvr = None  # Will be set when we encounter "NVR NUMBER" rows
    nvr_sections = []  # list of (nvr_number, start_col, header_row)

    # First pass: find NVR section headers
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            val = str(ws.cell_value(r, c)).strip().upper()
            if "NVR NUMBER" in val or "DVR NUMBER" in val:
                # Extract NVR number (e.g. "NVR NUMBER:- 1" -> 1)
                nums = re.findall(r'\d+', val)
                nvr_num = int(nums[0]) if nums else len(nvr_sections) + 1
                nvr_sections.append({"nvr": nvr_num, "col": c, "header_row": r})

    logger.info(f"Found {len(nvr_sections)} NVR sections: {nvr_sections}")

    # Second pass: for each NVR section, find the data header row and parse cameras
    for section in nvr_sections:
        start_col = section["col"]
        nvr_num = section["nvr"]
        header_row = section["header_row"]

        # Find the header row with "S.NO" / "CAMERA NAME" below the NVR header
        sno_col = None
        name_col = None
        loc_col = None
        sound_col = None

        for r in range(header_row, min(header_row + 5, ws.nrows)):
            for c in range(max(0, start_col - 1), min(ws.ncols, start_col + 5)):
                val = str(ws.cell_value(r, c)).strip().upper()
                if val in ("S.NO.", "S.NO", "SNO", "SR.NO"):
                    sno_col = c
                elif "CAMERA NAME" in val:
                    name_col = c
                elif "CAMERA LOCATION" in val or "LOCATION" in val:
                    loc_col = c
                elif "SOUND" in val:
                    sound_col = c

            if sno_col is not None and name_col is not None:
                data_start_row = r + 1
                break
        else:
            logger.warning(f"Could not find data headers for NVR {nvr_num}")
            continue

        # Parse camera rows
        for r in range(data_start_row, ws.nrows):
            sno_val = str(ws.cell_value(r, sno_col)).strip()
            if not sno_val or sno_val == "":
                # Check if we hit a new NVR section header
                row_text = " ".join(str(ws.cell_value(r, c)).strip() for c in range(max(0, start_col - 1), min(ws.ncols, start_col + 5)))
                if "NVR NUMBER" in row_text.upper() or "DVR NUMBER" in row_text.upper():
                    break
                continue

            # Extract channel number from S.NO
            try:
                channel = int(float(sno_val))
            except (ValueError, TypeError):
                continue

            cam_name = str(ws.cell_value(r, name_col)).strip() if name_col is not None else ""
            if not cam_name:
                continue

            location = str(ws.cell_value(r, loc_col)).strip() if loc_col is not None else ""
            sound = str(ws.cell_value(r, sound_col)).strip().upper() if sound_col is not None else ""

            mapping[cam_name] = {
                "dvr_index": nvr_num - 1,  # 0-based (NVR 1 -> index 0)
                "channel": channel,
                "location": location,
                "sound": sound == "YES",
                "description": f"{cam_name} ({location})",
            }

    return mapping


def _build_classroom_mapping(raw_mapping: dict) -> dict:
    """Convert raw camera mapping to classroom-focused mapping.

    Extracts classroom names from camera names like 'GRADE 10  CAM 2' -> 'Grade 10'.
    Groups multiple cameras per classroom and picks the best one (prefers CAM 1, with sound).
    """
    import re

    classroom_cameras = {}  # classroom -> list of camera entries

    for cam_name, info in raw_mapping.items():
        upper = cam_name.upper().strip()

        # Remove "CAM X" suffix first to isolate the classroom part
        classroom_part = re.sub(r'\s*CAM\s*\d+\s*$', '', upper).strip()
        # Collapse multiple spaces
        classroom_part = re.sub(r'\s+', ' ', classroom_part)

        # Check if this is a classroom camera
        match = re.match(
            r'((?:GRADE|NURSERY|NUR|PREP)\s+\d+\s*[A-C]?|(?:POPSICLES?|NURSERY|NUR|PREP))$',
            classroom_part,
        )

        if match:
            classroom = match.group(1).strip()
            # Normalize: "GRADE 10" -> "Grade 10", "NURSERY 3" -> "Nursery 3"
            classroom = classroom.title()
        else:
            # Not a classroom camera (library, sports room, etc.)
            # Still include it with the cleaned camera name as key
            classroom = classroom_part.title()

        if classroom not in classroom_cameras:
            classroom_cameras[classroom] = []
        classroom_cameras[classroom].append({
            "cam_name": cam_name,
            **info,
        })

    # Pick the best camera per classroom
    final_mapping = {}
    for classroom, cameras in classroom_cameras.items():
        # Prefer: CAM 1 > CAM 2, with sound > without sound
        best = sorted(cameras, key=lambda c: (
            "CAM 1" in c["cam_name"].upper(),  # True sorts after False, so negate
            c.get("sound", False),
        ), reverse=True)[0]

        final_mapping[classroom] = {
            "dvr_index": best["dvr_index"],
            "channel": best["channel"],
            "description": best["description"],
            "all_cameras": [
                {"name": c["cam_name"], "channel": c["channel"], "dvr_index": c["dvr_index"], "description": c.get("description", "")}
                for c in cameras
            ],
        }

    return final_mapping


@app.post("/api/mapping/upload")
async def upload_mapping(file: UploadFile = File(...)):
    """Upload an Excel file with camera-to-classroom mapping.

    Supports both .xls (old format) and .xlsx (new format).
    Auto-detects PPIS camera Excel layout with side-by-side NVR sections.
    """
    fname = file.filename or ""
    if not fname.endswith((".xlsx", ".xls")):
        return JSONResponse(
            {"status": "error", "error": "Please upload an .xlsx or .xls file"},
            status_code=400,
        )

    content = await file.read()
    ext = ".xls" if fname.endswith(".xls") else ".xlsx"
    upload_path = Path(__file__).parent / f"camera_mapping{ext}"
    with open(upload_path, "wb") as f:
        f.write(content)

    try:
        if ext == ".xls":
            # Use xlrd for old .xls format — auto-detect PPIS NVR layout
            raw_mapping = _parse_camera_xls(upload_path)
            mapping = _build_classroom_mapping(raw_mapping)
        else:
            # Use openpyxl for .xlsx
            import openpyxl
            wb = openpyxl.load_workbook(upload_path, data_only=True)
            ws = wb.active
            mapping = {}
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            classroom_col = None
            dvr_col = None
            channel_col = None
            desc_col = None

            for i, h in enumerate(headers):
                if "class" in h or "room" in h:
                    classroom_col = i
                elif "dvr" in h or "nvr" in h:
                    dvr_col = i
                elif "channel" in h or "camera" in h or "ch" == h:
                    channel_col = i
                elif "desc" in h or "note" in h or "label" in h:
                    desc_col = i

            if classroom_col is None or channel_col is None:
                # Fallback: try PPIS format detection on .xlsx too
                return JSONResponse(
                    {"status": "error", "error": "Excel must have columns: Classroom/Room and Channel/Camera. Or use the PPIS .xls format."},
                    status_code=400,
                )

            for row in ws.iter_rows(min_row=2, values_only=True):
                classroom = str(row[classroom_col] or "").strip()
                if not classroom:
                    continue
                dvr_num = int(row[dvr_col] or 1) if dvr_col is not None else 1
                channel = int(row[channel_col] or 1)
                desc = str(row[desc_col] or "") if desc_col is not None else ""
                mapping[classroom] = {
                    "dvr_index": dvr_num - 1,
                    "channel": channel,
                    "description": desc or classroom,
                }

        config["camera_mapping"] = mapping
        save_config(config)

        # Separate classroom cameras from non-classroom cameras for display
        classroom_keys = [k for k in mapping if any(
            kw in k.upper() for kw in ("GRADE", "NURSERY", "NUR", "PREP", "POPSICLE")
        )]
        other_keys = [k for k in mapping if k not in classroom_keys]

        return {
            "status": "ok",
            "mappings_loaded": len(mapping),
            "classroom_cameras": len(classroom_keys),
            "other_cameras": len(other_keys),
            "classrooms": sorted(classroom_keys),
            "other_locations": sorted(other_keys),
        }
    except Exception as e:
        logger.exception(f"Failed to parse camera mapping Excel: {e}")
        return JSONResponse(
            {"status": "error", "error": f"Failed to parse Excel: {str(e)}"},
            status_code=400,
        )


@app.post("/api/mapping/save")
async def save_mapping(request: Request):
    """Save camera mapping locally and sync to cloud."""
    body = await request.json()
    mapping = body.get("mapping", {})
    config["camera_mapping"] = mapping
    save_config(config)
    # Sync to cloud DB
    cloud_synced = False
    try:
        agent_secret = os.environ.get("AGENT_SECRET", "")
        headers = {"X-Agent-Secret": agent_secret} if agent_secret else {}
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                f"{CLOUD_API_BASE}/api/agent-config/camera-mapping",
                json={"camera_mapping": mapping},
                headers=headers,
            )
            cloud_synced = resp.status_code == 200
    except Exception as e:
        logger.warning(f"Failed to sync camera mapping to cloud: {e}")
    return {"status": "ok", "mappings_saved": len(mapping), "cloud_synced": cloud_synced}


@app.get("/api/snapshots")
async def list_snapshots():
    """List recent snapshots."""
    files = sorted(SNAPSHOT_DIR.glob("*.jpg"), key=os.path.getmtime, reverse=True)[:50]
    return [{"filename": f.name, "size": f.stat().st_size, "time": f.stat().st_mtime} for f in files]


# ---------------------------------------------------------------------------
# Face Recognition Attendance API
# ---------------------------------------------------------------------------

@app.post("/api/face/register")
async def register_face(
    image: UploadFile = File(...),
    person_id: str = Form(...),
    name: str = Form(...),
    role: str = Form(""),
    phone: str = Form(""),
    angle: str = Form("front"),
):
    """Register a face from an uploaded image."""
    image_bytes = await image.read()
    result = face_db.register_face(
        person_id=person_id,
        name=name,
        role=role,
        phone=phone,
        angle=angle,
        image_bytes=image_bytes,
    )
    if not result["success"]:
        return JSONResponse(result, status_code=400)
    # Reload known faces in the engine
    attendance_engine.reload_faces()
    # Sync to cloud DB
    cloud_synced = False
    try:
        agent_secret = os.environ.get("AGENT_SECRET", "")
        headers = {"X-Agent-Secret": agent_secret} if agent_secret else {}
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(
                f"{CLOUD_API_BASE}/api/face/register",
                data={"person_id": person_id, "name": name, "role": role, "phone": phone, "angle": angle},
                files={"image": ("face.jpg", image_bytes, "image/jpeg")},
                headers=headers,
            )
            cloud_synced = resp.status_code == 200
    except Exception as e:
        logger.warning(f"Failed to sync face to cloud: {e}")
    result["cloud_synced"] = cloud_synced
    return result


@app.get("/api/face/registered")
async def list_registered_faces():
    """List all registered persons."""
    return face_db.get_registered_list()


@app.put("/api/face/{person_id}/phone")
async def update_face_phone(person_id: str, request: Request):
    """Update the phone number for a registered face."""
    body = await request.json()
    new_phone = body.get("phone", "")
    if not new_phone:
        return {"status": "error", "error": "Missing phone"}
    import database as db_mod
    conn = db_mod.get_conn()
    try:
        cursor = conn.execute(
            "UPDATE registered_faces SET phone = ? WHERE person_id = ?",
            (new_phone, person_id),
        )
        conn.commit()
        updated = cursor.rowcount
    finally:
        conn.close()
    attendance_engine.reload_faces()
    return {"status": "ok", "updated": updated, "person_id": person_id, "phone": new_phone}


@app.delete("/api/face/{person_id}")
async def delete_registered_face(person_id: str):
    """Delete all face encodings for a person."""
    deleted = face_db.delete_person(person_id)
    attendance_engine.reload_faces()
    return {"deleted": deleted, "person_id": person_id}


@app.post("/api/face/migrate-insightface")
async def migrate_faces_to_insightface():
    """Re-encode all existing face photos with InsightFace (512-d ArcFace).

    Reads saved face images from disk and creates InsightFace embeddings
    alongside existing face_recognition encodings. Safe to run multiple
    times — skips faces that already have InsightFace encodings.
    """
    result = face_db.migrate_to_insightface()
    if result.get("success"):
        attendance_engine.reload_faces()
    return result


@app.post("/api/attendance/start")
async def start_attendance_monitoring(request: Request):
    """Start the face recognition attendance monitoring loop."""
    body = await request.json() if (request.headers.get("content-type") or "").startswith("application/json") else {}

    if attendance_engine.running:
        return {"status": "already_running"}
    if attendance_engine.classwise_running:
        return {"status": "error", "error": "Classwise monitoring is running. Stop it first."}

    # Configure engine
    attendance_engine.test_mode = body.get("test_mode", True)
    attendance_engine.test_person_id = body.get("test_person_id", "TEST001")
    attendance_engine.confidence_threshold = body.get("confidence_threshold", 0.30)
    attendance_engine.scan_interval = body.get("scan_interval", 3.0)
    attendance_engine.whatsapp_phone = body.get("whatsapp_phone", "")

    entrance = body.get("entrance_camera", None)
    dvrs = config.get("dvrs", [])

    attendance_engine.reload_faces()
    attendance_engine.running = True  # Set immediately to prevent duplicate starts
    attendance_engine._task = asyncio.create_task(
        attendance_engine.monitoring_loop(dvrs, entrance)
    )

    return {
        "status": "started",
        "config": attendance_engine.get_status(),
    }


@app.post("/api/attendance/start-classwise")
async def start_classwise_monitoring(request: Request):
    """Start classroom-wise face recognition on ALL classroom cameras.

    Each camera only checks students assigned to that class.
    Entry gate cameras check ALL registered faces.
    """
    body = await request.json() if (request.headers.get("content-type") or "").startswith("application/json") else {}

    if attendance_engine.classwise_running:
        return {"status": "already_running"}
    if attendance_engine.running:
        return {"status": "error", "error": "Single-camera monitoring is running. Stop it first."}

    attendance_engine.test_mode = False  # Classwise mode is always production
    attendance_engine.confidence_threshold = body.get("confidence_threshold", 0.30)
    attendance_engine.scan_interval = body.get("scan_interval", 3.0)

    dvrs = config.get("dvrs", [])
    camera_mapping = config.get("camera_mapping", {})

    if not dvrs:
        return {"status": "error", "error": "No DVRs configured"}
    if not camera_mapping:
        return {"status": "error", "error": "No camera mapping loaded"}

    cameras = attendance_engine.build_classroom_camera_list(camera_mapping, dvrs)
    classroom_cams = [c for c in cameras if c["grade"] is not None]
    gate_cams = [c for c in cameras if c["grade"] is None]

    attendance_engine.reload_faces()
    attendance_engine.classwise_running = True
    attendance_engine._classwise_task = asyncio.create_task(
        attendance_engine.classwise_monitoring_loop(dvrs, camera_mapping)
    )

    return {
        "status": "started",
        "classroom_cameras": len(classroom_cams),
        "gate_cameras": len(gate_cams),
        "total_faces": len(attendance_engine.known_faces),
        "grades_with_faces": len(attendance_engine._grade_face_cache),
        "config": attendance_engine.get_status(),
    }


@app.get("/api/attendance/classwise-cameras")
async def list_classwise_cameras():
    """List all classroom cameras with their grade assignments."""
    dvrs = config.get("dvrs", [])
    camera_mapping = config.get("camera_mapping", {})
    cameras = attendance_engine.build_classroom_camera_list(camera_mapping, dvrs)

    grade_face_counts = {
        g: len(v) for g, v in attendance_engine._grade_face_cache.items()
    }

    result = []
    for cam in cameras:
        grade = cam["grade"]
        result.append({
            "location": cam["location"],
            "grade": grade,
            "dvr_index": cam["dvr_index"],
            "channel": cam["channel"],
            "label": cam["label"],
            "faces_for_grade": grade_face_counts.get(grade, 0) if grade else len(attendance_engine.known_faces),
            "is_gate": cam.get("is_gate", grade is None),
        })

    return {
        "cameras": result,
        "total_classroom": sum(1 for c in result if not c["is_gate"]),
        "total_gate": sum(1 for c in result if c["is_gate"]),
        "grades_with_faces": grade_face_counts,
    }


@app.post("/api/attendance/stop")
async def stop_attendance_monitoring():
    """Stop attendance monitoring and disable auto-restart by the watchdog."""
    attendance_engine.stop()
    return {"status": "stopped", "auto_start_enabled": False}


# ---------------------------------------------------------------------------
# Mood Detection & Teacher Sighting Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/mood/status")
async def mood_status():
    """Get mood detection status."""
    return {
        "running": mood_detector.running,
        "tracked_persons": list(mood_detector._tracked_encodings.keys()),
        "cooldown_seconds": mood_detector._cooldown,
        "last_observations": {
            k: datetime.fromtimestamp(v).isoformat()
            for k, v in mood_detector._last_observation.items()
        } if mood_detector._last_observation else {},
    }


@app.get("/api/sighting/status")
async def sighting_status():
    """Get teacher sighting tracker + visitor tracking status."""
    return {
        "running": sighting_tracker.running,
        "teachers_loaded": len(sighting_tracker._teacher_encodings),
        "total_known_faces": len(sighting_tracker._all_encodings),
        "today": sighting_tracker._today,
        "sightings_today": len(sighting_tracker._daily_sightings),
        "visitors_today": len(sighting_tracker._daily_visitor_sightings),
        "recent_sightings": sighting_tracker._daily_sightings[-10:],
        "recent_visitors": sighting_tracker._daily_visitor_sightings[-10:],
    }


@app.post("/api/attendance/resend-notification")
async def resend_notification(person_id: str):
    """Clear notification dedup for a student so their next detection re-sends."""
    today = date.today().isoformat()
    cleared = []
    if attendance_engine._notification_sent.get(person_id) == today:
        del attendance_engine._notification_sent[person_id]
        cleared.append("notification_sent")
    if attendance_engine.daily_marked.get(person_id) == today:
        del attendance_engine.daily_marked[person_id]
        cleared.append("daily_marked")
    return {"person_id": person_id, "cleared": cleared}


@app.post("/api/attendance/scan-camera")
async def scan_specific_camera(location: str):
    """Manually trigger face recognition scan on a specific camera location."""
    mapping = attendance_engine._last_camera_mapping or {}
    dvrs = attendance_engine._last_dvrs or []
    # Find the camera
    target = location.strip()
    cam_data = None
    for key, val in mapping.items():
        if key.upper() == target.upper() or target.upper() in key.upper():
            cam_data = val
            target = key
            break
    if not cam_data:
        return {"error": f"Camera '{location}' not found", "available": list(mapping.keys())[:20]}
    dvr_idx = cam_data.get("dvr_index", 0)
    channel = cam_data.get("channel", 1)
    if dvr_idx >= len(dvrs):
        return {"error": f"DVR index {dvr_idx} out of range"}
    dvr = dvrs[dvr_idx]
    label = f"{target} (DVR {dvr_idx + 1} Ch {channel})"
    logger.info(f"Manual scan triggered for {label}")
    try:
        results = await attendance_engine.scan_camera(
            dvr, channel, label,
            faces_subset=None,  # Check ALL faces
            insightface_subset=None,
        )
        return {
            "camera": label,
            "faces_detected": len(results),
            "results": [
                {"person_id": r.get("person_id", ""), "name": r.get("name", ""),
                 "confidence": r.get("confidence", 0)}
                for r in results
            ] if results else [],
        }
    except Exception as e:
        return {"error": str(e), "camera": label}


@app.get("/api/attendance/status")
async def get_attendance_status():
    """Get attendance engine status."""
    return attendance_engine.get_status()


@app.get("/api/health")
async def health_check():
    """System health check — returns status of all subsystems."""
    health = attendance_engine._health.copy()
    health["classwise_running"] = attendance_engine.classwise_running
    health["single_cam_running"] = attendance_engine.running
    health["registered_faces"] = len(attendance_engine.known_faces)
    health["attendance_marked_today"] = sum(
        1 for d in attendance_engine.daily_marked.values()
        if d == __import__("datetime").date.today().isoformat()
    )
    health["cameras_with_errors"] = len(attendance_engine._camera_errors)

    # Memory info
    try:
        import psutil
        proc = psutil.Process()
        mem = proc.memory_info()
        health["memory_rss_mb"] = round(mem.rss / (1024 * 1024), 1)
        health["memory_vms_mb"] = round(mem.vms / (1024 * 1024), 1)
        health["cpu_percent"] = proc.cpu_percent(interval=0)
        health["open_files"] = len(proc.open_files())
        health["threads"] = proc.num_threads()
    except ImportError:
        pass
    except Exception:
        pass

    # Cache sizes
    health["sighting_cache_entries"] = len(attendance_engine._sightings)
    health["debug_log_entries"] = len(attendance_engine.debug_logs)
    health["background_tasks"] = len(attendance_engine._background_tasks)
    health["daily_marked_entries"] = len(attendance_engine.daily_marked)

    # Overall status
    statuses = [health["camera_feed"], health["recognition_engine"],
                health["notification_system"]]
    if "error" in statuses:
        health["overall"] = "error"
    elif "degraded" in statuses:
        health["overall"] = "degraded"
    else:
        health["overall"] = "ok"

    return health


@app.get("/api/attendance/logs")
async def get_attendance_logs(limit: int = 100, person_id: str | None = None):
    """Get attendance log entries from database."""
    import database as db_mod
    return db_mod.get_attendance_log(limit=limit, person_id=person_id)


@app.get("/api/attendance/debug")
async def get_debug_logs(limit: int = 100):
    """Get real-time debug logs from the attendance engine."""
    return attendance_engine.get_debug_logs(limit)


@app.post("/api/camera-alerts/configure")
async def configure_camera_alerts(request: Request):
    """Configure admin phones to receive camera offline/recovery alerts."""
    import database as db_mod
    body = await request.json()
    phones = body.get("phones", [])
    threshold = body.get("threshold", 5)
    if isinstance(phones, str):
        phones = [p.strip() for p in phones.split(",") if p.strip()]
    attendance_engine._admin_phones = phones
    attendance_engine._camera_alert_threshold = threshold
    # Persist to DB so it survives restarts
    db_mod.set_attendance_setting("camera_alert_phones", ",".join(phones))
    logger.info(f"Camera alerts configured: phones={phones}, threshold={threshold}")
    return {
        "status": "ok",
        "alert_phones": phones,
        "failure_threshold": threshold,
    }


@app.get("/api/camera-alerts/status")
async def camera_alerts_status():
    """Get current camera health and alert status."""
    errors = {}
    for cam_key, count in attendance_engine._camera_errors.items():
        label = attendance_engine._cam_key_to_label(cam_key)
        errors[label] = {"consecutive_failures": count, "alerted": cam_key in attendance_engine._admin_alerted}
    return {
        "alert_phones": attendance_engine._admin_phones,
        "failure_threshold": attendance_engine._camera_alert_threshold,
        "cameras_with_errors": errors,
        "total_alerted": len(attendance_engine._admin_alerted),
    }


@app.get("/api/camera-resolutions")
async def get_camera_resolutions():
    """Probe and report native resolution for all mapped cameras.

    Shows which cameras support HD (1080p), 2MP, 4MP, etc.
    This helps identify which cameras could benefit from replacement.
    """
    dvrs = config.get("dvrs", [])
    mapping = config.get("camera_mapping", {})
    if not dvrs or not mapping:
        return {"status": "error", "message": "No DVRs or camera mapping configured"}

    results = await attendance_engine.probe_all_camera_resolutions(dvrs, mapping)

    # Summarize
    total = len(results)
    hd_count = sum(1 for r in results.values() if r["width"] >= 1920)
    above_hd = sum(1 for r in results.values() if r["width"] > 1920)
    probed = sum(1 for r in results.values() if r["native_probed"])

    return {
        "status": "ok",
        "summary": {
            "total_cameras": total,
            "probed_successfully": probed,
            "hd_1080p_or_above": hd_count,
            "above_1080p": above_hd,
        },
        "cameras": results,
    }


@app.post("/api/attendance/sync-to-cloud")
async def sync_attendance_to_cloud():
    """Push all of today's attendance records to the cloud dashboard."""
    import database as db_mod
    from datetime import date, datetime, timedelta
    IST_OFFSET = timedelta(hours=5, minutes=30)
    records = db_mod.get_attendance_log(limit=500)
    today = date.today().isoformat()
    # Match both IST and UTC dates for today
    today_records = [r for r in records if today in str(r.get("logged_at", ""))]

    if not today_records:
        return {
            "status": "ok", "synced": 0,
            "message": "No attendance records today",
            "total_in_db": len(records),
            "today_str": today,
            "sample_dates": [str(r.get("logged_at", ""))[:16] for r in records[:5]],
        }

    api_url = os.environ.get("CLOUD_BOT_URL", "https://ppis-whatsapp-bot.fly.dev")
    agent_secret = os.environ.get("AGENT_SECRET", "")
    headers = {"Content-Type": "application/json"}
    if agent_secret:
        headers["X-Agent-Secret"] = agent_secret

    payload_records = []
    for r in today_records:
        pid = r.get("person_id", "")
        grade = ""
        for part in pid.split("_"):
            if part.startswith("GRADE") or part.startswith("NUR") or part.startswith("PREP"):
                grade = part
                break
        # Convert UTC DB time to IST for cloud display
        raw_time = str(r.get("logged_at", ""))
        try:
            dt = datetime.fromisoformat(raw_time)
            # If hour < 4, it's likely UTC — convert to IST
            if dt.hour < 4:
                dt = dt + IST_OFFSET
            raw_time = dt.isoformat()
        except Exception:
            pass
        payload_records.append({
            "person_id": pid,
            "name": r.get("name", ""),
            "grade": grade,
            "camera": r.get("camera_source", ""),
            "confidence": r.get("confidence", 0),
            "notification_sent": bool(r.get("whatsapp_sent")),
            "parent_phones": "",
            "logged_at": raw_time,
        })

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                f"{api_url}/api/dashboard/attendance/report",
                json={"records": payload_records},
                headers=headers,
            )
            data = resp.json()
            return {"status": "ok", "synced": data.get("inserted", 0), "total_today": len(today_records)}
    except Exception as e:
        logger.error(f"Cloud sync error: {type(e).__name__}: {e}")
        return {"status": "error", "error": f"{type(e).__name__}: {e}"}


@app.post("/api/attendance/reset-dedup")
async def reset_dedup():
    """Clear attendance dedup caches so students can be re-detected and notified."""
    marked_count = len(attendance_engine.daily_marked)
    notified_count = len(attendance_engine._notification_sent)
    attendance_engine.daily_marked.clear()
    attendance_engine._notification_sent.clear()
    return {
        "status": "ok",
        "cleared_marked": marked_count,
        "cleared_notified": notified_count,
        "message": "Dedup caches cleared — students will be re-detected and notified",
    }


@app.post("/api/attendance/retry-notifications")
async def retry_notifications():
    """Resend WhatsApp notifications for students marked today but not notified."""
    import database as db_mod
    from datetime import date, datetime, timedelta
    records = db_mod.get_attendance_log(limit=500)
    today = date.today().isoformat()
    today_records = [r for r in records if today in str(r.get("logged_at", ""))]

    # Filter to only those not yet notified
    not_notified = [r for r in today_records if not r.get("whatsapp_sent")]
    if not not_notified:
        return {"status": "ok", "message": "All students already notified", "total_today": len(today_records)}

    # IST offset (UTC+5:30)
    IST_OFFSET = timedelta(hours=5, minutes=30)

    sent_count = 0
    failed_count = 0
    for r in not_notified:
        pid = r.get("person_id", "")
        name = r.get("name", "")
        # Look up phone from face DB
        face_data = attendance_engine.known_faces.get(pid)
        if not face_data:
            continue
        phone = face_data.get("phone", "")
        if not phone:
            continue

        time_str = ""
        try:
            logged = str(r.get("logged_at", ""))
            dt = datetime.fromisoformat(logged)
            # DB stores UTC (datetime('now')), convert to IST
            dt_ist = dt + IST_OFFSET
            time_str = dt_ist.strftime("%I:%M %p")
        except Exception:
            time_str = "today"

        phone_list = [p.strip() for p in phone.split(",") if p.strip()]
        for parent_phone in phone_list:
            try:
                await attendance_engine._send_whatsapp_notification(
                    attendance_id=r.get("id", 0),
                    person_id=pid,
                    name=name,
                    time_str=time_str,
                    phone=parent_phone,
                )
                sent_count += 1
            except Exception as e:
                failed_count += 1
                logger.error(f"Retry notification failed for {name}: {e}")

    return {
        "status": "ok",
        "not_notified": len(not_notified),
        "attempted": sent_count + failed_count,
        "sent": sent_count,
        "failed": failed_count,
    }


@app.post("/api/attendance/resend-all")
async def resend_all_notifications():
    """Re-send attendance notifications to ALL parents with corrected IST times."""
    import database as db_mod
    from datetime import date, datetime, timedelta
    IST_OFFSET = timedelta(hours=5, minutes=30)

    records = db_mod.get_attendance_log(limit=500)
    today = date.today().isoformat()
    today_records = [r for r in records if today in str(r.get("logged_at", ""))]

    if not today_records:
        return {"status": "ok", "message": "No attendance records today"}

    # Skip staff/test faces
    skip_ids = {"arpit003", "Alisha002", "HARPREET001", "ALISHA001"}

    sent_count = 0
    failed_count = 0
    skipped_count = 0
    for r in today_records:
        pid = r.get("person_id", "")
        name = r.get("name", "")
        if pid in skip_ids:
            skipped_count += 1
            continue

        face_data = attendance_engine.known_faces.get(pid)
        if not face_data:
            skipped_count += 1
            continue
        phone = face_data.get("phone", "")
        if not phone:
            skipped_count += 1
            continue

        time_str = ""
        try:
            logged = str(r.get("logged_at", ""))
            dt = datetime.fromisoformat(logged)
            # Convert UTC to IST if hour < 4 (old UTC records)
            if dt.hour < 4:
                dt = dt + IST_OFFSET
            time_str = dt.strftime("%I:%M %p")
        except Exception:
            time_str = "today"

        phone_list = [p.strip() for p in phone.split(",") if p.strip()]
        for parent_phone in phone_list:
            try:
                await attendance_engine._send_whatsapp_notification(
                    attendance_id=r.get("id", 0),
                    person_id=pid,
                    name=name,
                    time_str=time_str,
                    phone=parent_phone,
                )
                sent_count += 1
            except Exception as e:
                failed_count += 1
                logger.error(f"Resend failed for {name}: {e}")

    return {
        "status": "ok",
        "total_today": len(today_records),
        "sent": sent_count,
        "failed": failed_count,
        "skipped": skipped_count,
    }


@app.post("/api/attendance/recognize")
async def recognize_single_image(image: UploadFile = File(...)):
    """Run face recognition on a single uploaded image (manual test)."""
    image_bytes = await image.read()
    attendance_engine.reload_faces()
    results = attendance_engine.recognize_faces_in_image(
        image_bytes, camera_source="manual_upload"
    )
    return {
        "results": results,
        "debug_logs": attendance_engine.get_debug_logs(20),
    }


@app.get("/api/attendance/today-summary")
async def get_today_summary():
    """Get today's attendance summary grouped by classroom."""
    import database as db_mod
    return {
        "total_marked": db_mod.get_today_attendance_count(),
        "by_classroom": db_mod.get_today_attendance_summary(),
    }


@app.get("/api/attendance/unrecognized")
async def get_unrecognized_faces(limit: int = 50, all: bool = False):
    """Get unrecognized faces flagged for manual review."""
    import database as db_mod
    return db_mod.get_unrecognized_faces(limit=limit, unreviewed_only=not all)


# ---------------------------------------------------------------------------
# A4 Sheet Capture — Controlled face registration
# ---------------------------------------------------------------------------

import a4_capture


@app.post("/api/a4-capture/single")
async def a4_capture_single(
    camera_label: str = Form(...),
    grade: str = Form(""),
):
    """Trigger A4 sheet capture for a single student on a specific camera.

    The student should be standing in front of the camera holding an A4 sheet
    with their name written in bold black text.
    """
    # Find the camera
    mapping = config.get("camera_mapping", {})
    cam_config = mapping.get(camera_label)
    if not cam_config:
        return JSONResponse({"success": False, "error": f"Camera '{camera_label}' not found"},
                           status_code=404)

    dvr_index = cam_config["dvr_index"]
    channel = cam_config["channel"]
    dvrs = config.get("dvrs", [])
    if dvr_index >= len(dvrs):
        return JSONResponse({"success": False, "error": "DVR not configured"},
                           status_code=400)

    dvr = dvrs[dvr_index]
    result = await a4_capture.capture_and_register(
        capture_func=capture_snapshot,
        dvr=dvr,
        channel=channel,
        camera_label=camera_label,
        grade=grade,
    )
    return result


@app.post("/api/a4-capture/batch")
async def a4_capture_batch(
    camera_label: str = Form(...),
    grade: str = Form(""),
    student_count: int = Form(1),
):
    """Trigger batch A4 sheet capture for multiple students in a class.

    Students line up and step in front of the camera one at a time.
    The system waits 5 seconds between each student.
    """
    mapping = config.get("camera_mapping", {})
    cam_config = mapping.get(camera_label)
    if not cam_config:
        return JSONResponse({"success": False, "error": f"Camera '{camera_label}' not found"},
                           status_code=404)

    dvr_index = cam_config["dvr_index"]
    channel = cam_config["channel"]
    dvrs = config.get("dvrs", [])
    if dvr_index >= len(dvrs):
        return JSONResponse({"success": False, "error": "DVR not configured"},
                           status_code=400)

    dvr = dvrs[dvr_index]
    results = await a4_capture.batch_capture_class(
        capture_func=capture_snapshot,
        dvr=dvr,
        channel=channel,
        camera_label=camera_label,
        grade=grade,
        student_count=student_count,
    )

    success_count = sum(1 for r in results if r.get("success"))
    return {
        "total": student_count,
        "success": success_count,
        "failed": student_count - success_count,
        "results": results,
    }


@app.get("/api/a4-capture/logs")
async def get_a4_capture_logs(date: str = ""):
    """Get A4 capture logs for a given date (default: today IST)."""
    logs = a4_capture.get_capture_logs(date)
    return {"date": date, "entries": logs}


@app.get("/api/a4-capture/cameras")
async def get_capture_cameras():
    """Get list of cameras available for A4 sheet capture."""
    mapping = config.get("camera_mapping", {})
    cameras = []
    for label, cam_config in mapping.items():
        cameras.append({
            "label": label,
            "dvr_index": cam_config.get("dvr_index"),
            "channel": cam_config.get("channel"),
            "description": cam_config.get("description", ""),
        })
    return {"cameras": cameras}


# ---------------------------------------------------------------------------
# Dashboard HTML (embedded for simplicity — no external templates needed)
# ---------------------------------------------------------------------------

def get_dashboard_html() -> str:
    return """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PPIS Campus Agent</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', Tahoma, sans-serif; background: #f0f2f5; color: #333; }
.header { background: linear-gradient(135deg, #1a237e, #283593); color: white; padding: 20px 30px; display: flex; align-items: center; gap: 15px; }
.header h1 { font-size: 24px; font-weight: 600; }
.header .status { margin-left: auto; padding: 6px 16px; border-radius: 20px; font-size: 13px; font-weight: 500; }
.status.connected { background: #43a047; }
.status.disconnected { background: #e53935; }
.container { max-width: 1200px; margin: 20px auto; padding: 0 20px; }
.tabs { display: flex; gap: 4px; margin-bottom: 20px; }
.tab { padding: 10px 24px; background: white; border: 1px solid #ddd; border-radius: 8px 8px 0 0; cursor: pointer; font-weight: 500; }
.tab.active { background: #1a237e; color: white; border-color: #1a237e; }
.panel { display: none; background: white; border-radius: 0 8px 8px 8px; padding: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
.panel.active { display: block; }
.card { border: 1px solid #e0e0e0; border-radius: 8px; padding: 16px; margin-bottom: 16px; }
.card h3 { margin-bottom: 12px; color: #1a237e; }
.form-row { display: flex; gap: 12px; margin-bottom: 12px; align-items: center; }
.form-row label { min-width: 100px; font-weight: 500; }
.form-row input, .form-row select { flex: 1; padding: 8px 12px; border: 1px solid #ccc; border-radius: 6px; font-size: 14px; }
button { padding: 8px 20px; background: #1a237e; color: white; border: none; border-radius: 6px; cursor: pointer; font-size: 14px; font-weight: 500; }
button:hover { background: #283593; }
button.test { background: #43a047; }
button.test:hover { background: #388e3c; }
button.danger { background: #e53935; }
.mapping-table { width: 100%; border-collapse: collapse; margin-top: 12px; }
.mapping-table th, .mapping-table td { padding: 8px 12px; border: 1px solid #e0e0e0; text-align: left; }
.mapping-table th { background: #f5f5f5; font-weight: 600; }
.upload-zone { border: 2px dashed #ccc; border-radius: 8px; padding: 40px; text-align: center; cursor: pointer; transition: all 0.2s; }
.upload-zone:hover { border-color: #1a237e; background: #f8f9ff; }
.snapshot-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; margin-top: 12px; }
.snapshot-card { border: 1px solid #e0e0e0; border-radius: 8px; overflow: hidden; }
.snapshot-card img { width: 100%; height: 150px; object-fit: cover; }
.snapshot-card .info { padding: 8px; font-size: 12px; color: #666; }
.alert { padding: 12px 16px; border-radius: 6px; margin-bottom: 12px; }
.alert.success { background: #e8f5e9; color: #2e7d32; border: 1px solid #a5d6a7; }
.alert.error { background: #ffebee; color: #c62828; border: 1px solid #ef9a9a; }
.alert.info { background: #e3f2fd; color: #1565c0; border: 1px solid #90caf9; }
#alert-box { position: fixed; top: 20px; right: 20px; z-index: 1000; min-width: 300px; }
</style>
</head>
<body>

<div class="header">
    <div>🏫</div>
    <h1>PPIS Campus Agent</h1>
    <div class="status disconnected" id="ws-status">Disconnected</div>
</div>

<div id="alert-box"></div>

<div class="container">
    <div class="tabs">
        <div class="tab active" onclick="switchTab('attendance')">Attendance</div>
        <div class="tab" onclick="switchTab('register')">Face Registration</div>
        <div class="tab" onclick="switchTab('dvr')">DVR Configuration</div>
        <div class="tab" onclick="switchTab('mapping')">Camera Mapping</div>
        <div class="tab" onclick="switchTab('snapshots')">Snapshots</div>
        <div class="tab" onclick="switchTab('logs')">Logs</div>
    </div>

    <!-- Attendance Panel -->
    <div class="panel active" id="panel-attendance">
        <h2 style="margin-bottom:16px">Face Recognition Attendance</h2>

        <div class="card">
            <h3>Monitoring Controls</h3>
            <div class="form-row">
                <label>Status:</label>
                <span id="att-status" style="font-weight:bold;color:#e53935">Stopped</span>
            </div>
            <div class="form-row">
                <label>Test Mode:</label>
                <select id="att-test-mode"><option value="true" selected>ON (TEST001 only)</option><option value="false">OFF (All persons)</option></select>
            </div>
            <div class="form-row">
                <label>Test Person ID:</label>
                <input type="text" id="att-test-pid" value="TEST001">
            </div>
            <div class="form-row">
                <label>Confidence:</label>
                <input type="number" id="att-threshold" value="85" min="50" max="100" step="1">%
            </div>
            <div class="form-row">
                <label>Scan Interval:</label>
                <input type="number" id="att-interval" value="3" min="1" max="30" step="1">s
            </div>
            <div class="form-row">
                <label>WhatsApp #:</label>
                <input type="text" id="att-phone" placeholder="e.g. +91XXXXXXXXXX">
            </div>
            <div class="form-row">
                <label>DVR Index:</label>
                <input type="number" id="att-dvr-idx" value="0" min="0">
            </div>
            <div class="form-row">
                <label>Channel:</label>
                <input type="number" id="att-channel" value="1" min="1">
            </div>
            <button onclick="startAttendance()" style="background:#43a047">Start Monitoring</button>
            <button onclick="stopAttendance()" class="danger" style="margin-left:8px">Stop</button>
            <button onclick="refreshAttStatus()" style="margin-left:8px">Refresh Status</button>
        </div>

        <div class="card" style="border-left:4px solid #1565c0;background:#e3f2fd">
            <h3 style="color:#1565c0">Classroom-wise Attendance (All Cameras)</h3>
            <p style="margin-bottom:12px;color:#555">
                Scan ALL classroom cameras simultaneously. Each camera only checks students assigned to that class.
                Entry gate cameras check all registered faces.
            </p>
            <div class="form-row">
                <label>Confidence:</label>
                <input type="number" id="cw-threshold" value="40" min="20" max="100" step="1">%
            </div>
            <button onclick="startClasswise()" style="background:#1565c0;color:white">
                Start All Classrooms
            </button>
            <button onclick="stopAttendance()" class="danger" style="margin-left:8px">Stop All</button>
            <button onclick="loadClasswiseCameras()" style="margin-left:8px">View Camera List</button>
            <div id="classwise-status" style="margin-top:12px;font-family:monospace;font-size:13px;white-space:pre-wrap;background:#fff;padding:12px;border-radius:6px;display:none"></div>
            <div id="classwise-cameras" style="margin-top:12px;display:none">
                <table class="mapping-table">
                    <thead><tr><th>Location</th><th>Grade</th><th>DVR</th><th>Channel</th><th>Faces</th><th>Type</th></tr></thead>
                    <tbody id="cw-cameras-body"></tbody>
                </table>
            </div>
        </div>

        <div class="card">
            <h3>Manual Test - Upload Image</h3>
            <p style="margin-bottom:12px;color:#666">Upload an image to test face recognition without live camera feed.</p>
            <input type="file" id="test-image-upload" accept="image/*">
            <button onclick="testRecognize()" style="margin-top:8px">Recognize Faces</button>
            <div id="recognize-result" style="margin-top:12px;font-family:monospace;font-size:13px;white-space:pre-wrap"></div>
        </div>

        <div class="card">
            <h3>Attendance Log</h3>
            <button onclick="loadAttendanceLogs()">Refresh Logs</button>
            <table class="mapping-table" style="margin-top:12px" id="att-log-table">
                <thead><tr><th>Name</th><th>ID</th><th>Time</th><th>Status</th><th>Confidence</th><th>Camera</th><th>WhatsApp</th></tr></thead>
                <tbody id="att-log-body"></tbody>
            </table>
        </div>

        <div class="card">
            <h3>Debug Logs</h3>
            <button onclick="loadDebugLogs()">Refresh</button>
            <div id="debug-log-container" style="background:#1e1e1e;color:#d4d4d4;padding:16px;border-radius:8px;height:300px;overflow-y:auto;font-family:monospace;font-size:12px;margin-top:12px">
                <p>No debug logs yet. Start monitoring to see real-time events.</p>
            </div>
        </div>
    </div>

    <!-- Face Registration Panel -->
    <div class="panel" id="panel-register">
        <h2 style="margin-bottom:16px">Face Registration</h2>

        <div class="card">
            <h3>Register New Face</h3>
            <p style="margin-bottom:12px;color:#666">Upload face images from multiple angles (front, left, right) for better recognition accuracy.</p>
            <div class="form-row"><label>Person ID:</label><input type="text" id="reg-pid" value="TEST001" placeholder="e.g. TEST001"></div>
            <div class="form-row"><label>Name:</label><input type="text" id="reg-name" placeholder="Your Name"></div>
            <div class="form-row"><label>Role:</label><input type="text" id="reg-role" value="Test User" placeholder="e.g. Student, Teacher"></div>
            <div class="form-row"><label>Phone:</label><input type="text" id="reg-phone" placeholder="+91XXXXXXXXXX"></div>
            <div class="form-row"><label>Angle:</label>
                <select id="reg-angle">
                    <option value="front">Front</option>
                    <option value="left">Left</option>
                    <option value="right">Right</option>
                </select>
            </div>
            <div class="form-row"><label>Image:</label><input type="file" id="reg-image" accept="image/*"></div>
            <button onclick="registerFace()">Register Face</button>
            <div id="reg-result" style="margin-top:12px"></div>
        </div>

        <div class="card">
            <h3>Registered Persons</h3>
            <button onclick="loadRegistered()">Refresh</button>
            <table class="mapping-table" style="margin-top:12px">
                <thead><tr><th>Person ID</th><th>Name</th><th>Role</th><th>Phone</th><th>Faces</th><th>Angles</th><th>Action</th></tr></thead>
                <tbody id="registered-body"></tbody>
            </table>
        </div>
    </div>

    <!-- DVR Configuration Panel -->
    <div class="panel" id="panel-dvr">
        <h2 style="margin-bottom:16px">DVR Configuration</h2>
        <div id="dvr-list"></div>
        <button onclick="addDvr()" style="margin-top:12px">+ Add DVR</button>
        <button onclick="saveDvrs()" style="margin-top:12px; margin-left:8px">Save Configuration</button>
    </div>

    <!-- Camera Mapping Panel -->
    <div class="panel" id="panel-mapping">
        <h2 style="margin-bottom:16px">Camera-to-Classroom Mapping</h2>

        <div class="card">
            <h3>Upload Excel Mapping</h3>
            <p style="margin-bottom:12px; color:#666">Upload an Excel file with columns: <b>Classroom</b>, <b>DVR</b> (1/2/3), <b>Channel</b>, <b>Description</b> (optional)</p>
            <div class="upload-zone" onclick="document.getElementById('excel-upload').click()">
                <p>📁 Click to upload Excel file (.xlsx)</p>
                <p style="font-size:12px; color:#999; margin-top:8px">or drag and drop here</p>
            </div>
            <input type="file" id="excel-upload" accept=".xlsx,.xls" style="display:none" onchange="uploadExcel(this)">
        </div>

        <div class="card">
            <h3>Current Mapping</h3>
            <div id="mapping-table-container">
                <p style="color:#999">No mapping loaded. Upload an Excel file above.</p>
            </div>
        </div>

        <div class="card">
            <h3>Add Manual Entry</h3>
            <div class="form-row">
                <label>Classroom:</label>
                <input type="text" id="map-classroom" placeholder="e.g. Grade 3C">
            </div>
            <div class="form-row">
                <label>DVR #:</label>
                <select id="map-dvr"><option value="1">DVR 1</option><option value="2">DVR 2</option><option value="3">DVR 3</option></select>
            </div>
            <div class="form-row">
                <label>Channel:</label>
                <input type="number" id="map-channel" min="1" max="64" value="1">
            </div>
            <div class="form-row">
                <label>Description:</label>
                <input type="text" id="map-desc" placeholder="Optional description">
            </div>
            <button onclick="addMapping()">Add Mapping</button>
        </div>
    </div>

    <!-- Snapshots Panel -->
    <div class="panel" id="panel-snapshots">
        <h2 style="margin-bottom:16px">Snapshots</h2>
        <div class="form-row">
            <label>Classroom:</label>
            <input type="text" id="snap-classroom" placeholder="e.g. Grade 3C">
            <button onclick="takeSnapshot()">📷 Capture Snapshot</button>
        </div>
        <div class="snapshot-grid" id="snapshot-grid"></div>
    </div>

    <!-- Logs Panel -->
    <div class="panel" id="panel-logs">
        <h2 style="margin-bottom:16px">Activity Logs</h2>
        <div id="log-container" style="background:#1e1e1e; color:#d4d4d4; padding:16px; border-radius:8px; height:400px; overflow-y:auto; font-family:monospace; font-size:13px;">
            <p>Agent started. Waiting for events...</p>
        </div>
    </div>
</div>

<script>
let currentMapping = {};
let dvrs = [];

function switchTab(tab) {
    document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
    event.target.classList.add('active');
    document.getElementById('panel-' + tab).classList.add('active');
    if (tab === 'snapshots') loadSnapshots();
    if (tab === 'attendance') { refreshAttStatus(); loadAttendanceLogs(); }
    if (tab === 'register') loadRegistered();
}

function showAlert(msg, type) {
    const box = document.getElementById('alert-box');
    const el = document.createElement('div');
    el.className = 'alert ' + type;
    el.textContent = msg;
    box.appendChild(el);
    setTimeout(() => el.remove(), 5000);
}

function addLog(msg) {
    const container = document.getElementById('log-container');
    const ts = new Date().toLocaleTimeString();
    container.innerHTML += `<p>[${ts}] ${msg}</p>`;
    container.scrollTop = container.scrollHeight;
}

// --- DVR ---
function renderDvrs() {
    const list = document.getElementById('dvr-list');
    list.innerHTML = '';
    dvrs.forEach((d, i) => {
        list.innerHTML += `
        <div class="card">
            <h3>${d.name || 'DVR ' + (i+1)}</h3>
            <div class="form-row"><label>Name:</label><input value="${d.name||''}" onchange="dvrs[${i}].name=this.value"></div>
            <div class="form-row"><label>IP Address:</label><input value="${d.ip||''}" onchange="dvrs[${i}].ip=this.value"></div>
            <div class="form-row"><label>Port:</label><input type="number" value="${d.port||80}" onchange="dvrs[${i}].port=parseInt(this.value)"></div>
            <div class="form-row"><label>Username:</label><input value="${d.username||''}" onchange="dvrs[${i}].username=this.value"></div>
            <div class="form-row"><label>Password:</label><input type="password" value="${d.password||''}" onchange="dvrs[${i}].password=this.value"></div>
            <div class="form-row"><label>Channels:</label><input type="number" value="${d.channels||64}" onchange="dvrs[${i}].channels=parseInt(this.value)"></div>
            <button class="test" onclick="testDvr(${i})">Test Connection</button>
            <button class="danger" onclick="dvrs.splice(${i},1);renderDvrs()" style="margin-left:8px">Remove</button>
            <span id="dvr-test-${i}" style="margin-left:12px;font-size:13px"></span>
        </div>`;
    });
}

function addDvr() {
    dvrs.push({name:'DVR '+(dvrs.length+1), ip:'192.168.0.11', port:80, username:'admin', password:'', channels:64});
    renderDvrs();
}

async function saveDvrs() {
    const resp = await fetch('/api/dvr/save', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({dvrs})});
    const data = await resp.json();
    showAlert('DVR configuration saved (' + data.dvr_count + ' DVRs)', 'success');
    addLog('DVR configuration saved');
}

async function testDvr(i) {
    const el = document.getElementById('dvr-test-' + i);
    el.textContent = 'Testing...';
    el.style.color = '#666';
    // Need to save first so backend has the password
    await fetch('/api/dvr/save', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({dvrs})});
    const resp = await fetch('/api/dvr/test/' + i, {method:'POST'});
    const data = await resp.json();
    if (data.status === 'connected') {
        el.textContent = '✓ Connected';
        el.style.color = 'green';
        addLog('DVR ' + (i+1) + ' (' + dvrs[i].ip + '): Connected');
    } else {
        el.textContent = '✗ ' + (data.error || data.status);
        el.style.color = 'red';
        addLog('DVR ' + (i+1) + ' (' + dvrs[i].ip + '): ' + data.error);
    }
}

// --- Mapping ---
function renderMapping() {
    const container = document.getElementById('mapping-table-container');
    const keys = Object.keys(currentMapping);
    if (keys.length === 0) {
        container.innerHTML = '<p style="color:#999">No mapping loaded.</p>';
        return;
    }
    let html = '<table class="mapping-table"><thead><tr><th>Classroom</th><th>DVR</th><th>Channel</th><th>Description</th><th>Action</th></tr></thead><tbody>';
    keys.forEach(k => {
        const m = currentMapping[k];
        html += `<tr><td>${k}</td><td>DVR ${(m.dvr_index||0)+1}</td><td>${m.channel}</td><td>${m.description||''}</td><td><button class="danger" onclick="deleteMapping('${k}')">Delete</button></td></tr>`;
    });
    html += '</tbody></table>';
    container.innerHTML = html;
}

function addMapping() {
    const classroom = document.getElementById('map-classroom').value.trim();
    const dvrNum = parseInt(document.getElementById('map-dvr').value);
    const channel = parseInt(document.getElementById('map-channel').value);
    const desc = document.getElementById('map-desc').value.trim();
    if (!classroom) { showAlert('Please enter a classroom name', 'error'); return; }
    currentMapping[classroom] = {dvr_index: dvrNum-1, channel, description: desc || classroom};
    renderMapping();
    saveMapping();
    document.getElementById('map-classroom').value = '';
    document.getElementById('map-desc').value = '';
    showAlert('Mapping added: ' + classroom + ' → DVR ' + dvrNum + ' Ch' + channel, 'success');
}

async function deleteMapping(key) {
    delete currentMapping[key];
    renderMapping();
    await saveMapping();
    showAlert('Mapping removed: ' + key, 'info');
}

async function saveMapping() {
    await fetch('/api/mapping/save', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({mapping: currentMapping})});
}

async function uploadExcel(input) {
    const file = input.files[0];
    if (!file) return;
    const form = new FormData();
    form.append('file', file);
    const resp = await fetch('/api/mapping/upload', {method:'POST', body: form});
    const data = await resp.json();
    if (data.status === 'ok') {
        showAlert('Loaded ' + data.mappings_loaded + ' camera mappings from Excel', 'success');
        addLog('Camera mapping uploaded: ' + data.mappings_loaded + ' entries');
        await loadConfig();
    } else {
        showAlert('Error: ' + data.error, 'error');
    }
}

// --- Snapshots ---
async function takeSnapshot() {
    const classroom = document.getElementById('snap-classroom').value.trim();
    if (!classroom) { showAlert('Enter a classroom name', 'error'); return; }
    showAlert('Capturing snapshot for ' + classroom + '...', 'info');
    const resp = await fetch('/api/snapshot/' + encodeURIComponent(classroom), {method:'POST'});
    const data = await resp.json();
    if (data.status === 'ok') {
        showAlert('Snapshot captured: ' + data.filename, 'success');
        addLog('Snapshot: ' + classroom + ' → ' + data.filename);
        loadSnapshots();
    } else {
        showAlert('Error: ' + data.error, 'error');
    }
}

async function loadSnapshots() {
    const resp = await fetch('/api/snapshots');
    const data = await resp.json();
    const grid = document.getElementById('snapshot-grid');
    grid.innerHTML = '';
    data.forEach(s => {
        const date = new Date(s.time * 1000).toLocaleString();
        grid.innerHTML += `<div class="snapshot-card"><img src="/snapshots/${s.filename}" alt="${s.filename}"><div class="info">${s.filename}<br>${date}<br>${Math.round(s.size/1024)}KB</div></div>`;
    });
}

// --- Face Registration ---
async function registerFace() {
    const pid = document.getElementById('reg-pid').value.trim();
    const name = document.getElementById('reg-name').value.trim();
    const role = document.getElementById('reg-role').value.trim();
    const phone = document.getElementById('reg-phone').value.trim();
    const angle = document.getElementById('reg-angle').value;
    const fileInput = document.getElementById('reg-image');
    if (!pid || !name) { showAlert('Person ID and Name are required', 'error'); return; }
    if (!fileInput.files[0]) { showAlert('Please select an image', 'error'); return; }
    const form = new FormData();
    form.append('person_id', pid);
    form.append('name', name);
    form.append('role', role);
    form.append('phone', phone);
    form.append('angle', angle);
    form.append('image', fileInput.files[0]);
    const resp = await fetch('/api/face/register', {method:'POST', body: form});
    const data = await resp.json();
    const el = document.getElementById('reg-result');
    if (data.success) {
        el.innerHTML = '<div class="alert success">Face registered: ' + name + ' (' + angle + ')</div>';
        showAlert('Face registered for ' + name + ' (' + angle + ')', 'success');
        loadRegistered();
    } else {
        el.innerHTML = '<div class="alert error">Error: ' + data.error + '</div>';
        showAlert('Registration failed: ' + data.error, 'error');
    }
}

async function loadRegistered() {
    const resp = await fetch('/api/face/registered');
    const data = await resp.json();
    const tbody = document.getElementById('registered-body');
    tbody.innerHTML = '';
    data.forEach(p => {
        tbody.innerHTML += '<tr><td>' + p.person_id + '</td><td>' + p.name + '</td><td>' + p.role + '</td><td>' + p.phone + '</td><td>' + p.face_count + '</td><td>' + (p.angles||'') + '</td><td><button class="danger" onclick="deletePerson(\\'' + p.person_id + '\\')">Delete</button></td></tr>';
    });
}

async function deletePerson(pid) {
    if (!confirm('Delete all face data for ' + pid + '?')) return;
    await fetch('/api/face/' + pid, {method:'DELETE'});
    showAlert('Deleted face data for ' + pid, 'info');
    loadRegistered();
}

// --- Attendance Monitoring ---
async function startAttendance() {
    const body = {
        test_mode: document.getElementById('att-test-mode').value === 'true',
        test_person_id: document.getElementById('att-test-pid').value.trim(),
        confidence_threshold: parseInt(document.getElementById('att-threshold').value) / 100.0,
        scan_interval: parseFloat(document.getElementById('att-interval').value),
        whatsapp_phone: document.getElementById('att-phone').value.trim(),
        entrance_camera: {
            dvr_index: parseInt(document.getElementById('att-dvr-idx').value),
            channel: parseInt(document.getElementById('att-channel').value),
            label: 'Entrance'
        }
    };
    const resp = await fetch('/api/attendance/start', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    const data = await resp.json();
    showAlert('Attendance monitoring: ' + data.status, data.status === 'started' ? 'success' : 'info');
    refreshAttStatus();
}

async function stopAttendance() {
    await fetch('/api/attendance/stop', {method:'POST'});
    showAlert('Attendance monitoring stopped', 'info');
    refreshAttStatus();
}

async function refreshAttStatus() {
    try {
        const resp = await fetch('/api/attendance/status');
        const data = await resp.json();
        const el = document.getElementById('att-status');
        if (data.classwise_running) {
            el.textContent = 'CLASSWISE RUNNING (' + data.registered_persons + ' persons, ' + data.grades_with_faces + ' grades, ' + data.attendance_marked_today + ' marked today)';
            el.style.color = '#1565c0';
            // Update classwise status panel
            const cwEl = document.getElementById('classwise-status');
            cwEl.style.display = 'block';
            const s = data.classwise_stats || {};
            cwEl.innerHTML = '<b>Status:</b> RUNNING<br>' +
                '<b>Cameras:</b> ' + s.total_cameras + '<br>' +
                '<b>Current:</b> ' + (s.current_camera || '-') + '<br>' +
                '<b>Cycle:</b> #' + s.cycle_count + ' (' + s.last_cycle_duration + 's)<br>' +
                '<b>Faces detected:</b> ' + s.faces_detected_total + '<br>' +
                '<b>Attendance today:</b> ' + s.attendance_marked_today + '<br>' +
                '<b>Errors:</b> ' + s.errors;
        } else if (data.running) {
            el.textContent = 'Running (' + data.registered_persons + ' persons, ' + data.total_encodings + ' encodings)';
            el.style.color = '#43a047';
            document.getElementById('classwise-status').style.display = 'none';
        } else {
            el.textContent = 'Stopped';
            el.style.color = '#e53935';
            document.getElementById('classwise-status').style.display = 'none';
        }
    } catch(e) {}
}

async function loadAttendanceLogs() {
    const resp = await fetch('/api/attendance/logs?limit=50');
    const data = await resp.json();
    const tbody = document.getElementById('att-log-body');
    tbody.innerHTML = '';
    data.forEach(l => {
        tbody.innerHTML += '<tr><td>' + l.name + '</td><td>' + l.person_id + '</td><td>' + l.logged_at + '</td><td>' + l.status + '</td><td>' + (l.confidence * 100).toFixed(1) + '%</td><td>' + l.camera_source + '</td><td>' + (l.whatsapp_sent ? 'Sent' : '-') + '</td></tr>';
    });
}

async function loadDebugLogs() {
    const resp = await fetch('/api/attendance/debug?limit=100');
    const data = await resp.json();
    const container = document.getElementById('debug-log-container');
    container.innerHTML = '';
    data.forEach(l => {
        let color = '#d4d4d4';
        if (l.event === 'face_matched' || l.event === 'attendance_marked') color = '#4caf50';
        else if (l.event === 'error' || l.event === 'whatsapp_error') color = '#ef5350';
        else if (l.event === 'face_detected') color = '#42a5f5';
        else if (l.event === 'low_confidence') color = '#ffa726';
        container.innerHTML += '<p style="color:' + color + '">[' + l.timestamp + '] <b>' + l.event + '</b>: ' + l.details + (l.person_id ? ' (ID: ' + l.person_id + ')' : '') + (l.confidence > 0 ? ' [' + (l.confidence * 100).toFixed(1) + '%]' : '') + '</p>';
    });
    container.scrollTop = container.scrollHeight;
}

async function testRecognize() {
    const fileInput = document.getElementById('test-image-upload');
    if (!fileInput.files[0]) { showAlert('Select an image first', 'error'); return; }
    const form = new FormData();
    form.append('image', fileInput.files[0]);
    const el = document.getElementById('recognize-result');
    el.textContent = 'Processing...';
    const resp = await fetch('/api/attendance/recognize', {method:'POST', body: form});
    const data = await resp.json();
    let output = '=== Recognition Results ===\\n';
    if (data.results && data.results.length > 0) {
        data.results.forEach(r => {
            output += 'MATCH: ' + r.name + ' (ID: ' + r.person_id + ') - Confidence: ' + (r.confidence * 100).toFixed(1) + '% - Status: ' + r.status + '\\n';
        });
    } else {
        output += 'No matching faces found.\\n';
    }
    output += '\\n=== Debug Logs ===\\n';
    (data.debug_logs || []).forEach(l => {
        output += '[' + l.event + '] ' + l.details + '\\n';
    });
    el.textContent = output;
}

// --- Classwise Monitoring ---
async function startClasswise() {
    const body = {
        confidence_threshold: parseInt(document.getElementById('cw-threshold').value) / 100.0,
    };
    const resp = await fetch('/api/attendance/start-classwise', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    const data = await resp.json();
    if (data.status === 'started') {
        showAlert('Classwise monitoring started: ' + data.classroom_cameras + ' classrooms + ' + data.gate_cameras + ' gates, ' + data.total_faces + ' faces', 'success');
    } else {
        showAlert('Classwise: ' + (data.error || data.status), 'error');
    }
    refreshAttStatus();
}

async function loadClasswiseCameras() {
    const resp = await fetch('/api/attendance/classwise-cameras');
    const data = await resp.json();
    const tbody = document.getElementById('cw-cameras-body');
    tbody.innerHTML = '';
    (data.cameras || []).forEach(c => {
        tbody.innerHTML += '<tr><td>' + c.location + '</td><td>' + (c.grade || 'ALL') + '</td><td>DVR ' + (c.dvr_index + 1) + '</td><td>' + c.channel + '</td><td>' + c.faces_for_grade + '</td><td>' + (c.is_gate ? 'Entry Gate' : 'Classroom') + '</td></tr>';
    });
    document.getElementById('classwise-cameras').style.display = 'block';
    showAlert('Found ' + data.total_classroom + ' classroom cameras + ' + data.total_gate + ' entry gate cameras', 'info');
}

// --- Init ---
async function loadConfig() {
    const resp = await fetch('/api/config');
    const data = await resp.json();
    dvrs = data.dvrs.map(d => ({...d, password: ''}));
    currentMapping = data.camera_mapping || {};
    renderDvrs();
    renderMapping();
    const statusEl = document.getElementById('ws-status');
    if (data.ws_connected) {
        statusEl.textContent = 'Connected to Cloud Bot';
        statusEl.className = 'status connected';
    } else {
        statusEl.textContent = 'Disconnected';
        statusEl.className = 'status disconnected';
    }
}

// Load config, attendance status, and refresh periodically
loadConfig();
refreshAttStatus();
loadAttendanceLogs();
loadRegistered();
setInterval(async () => {
    try {
        const resp = await fetch('/api/config');
        const data = await resp.json();
        const statusEl = document.getElementById('ws-status');
        if (data.ws_connected) {
            statusEl.textContent = 'Connected to Cloud Bot';
            statusEl.className = 'status connected';
        } else {
            statusEl.textContent = 'Disconnected';
            statusEl.className = 'status disconnected';
        }
    } catch(e) {}
    // Auto-refresh attendance status
    refreshAttStatus();
}, 5000);
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


if __name__ == "__main__":
    import uvicorn
    import traceback

    # Global crash logger
    def _log_crash(exc_type, exc_value, exc_tb):
        logger.critical(
            f"UNHANDLED EXCEPTION: {exc_type.__name__}: {exc_value}\n"
            + "".join(traceback.format_tb(exc_tb))
        )
    sys.excepthook = _log_crash

    port = config.get("local_port", 8897)

    logger.info(f"Starting PPIS Campus Agent on http://localhost:{port}")

    # Retry binding up to 5 times with delays — handles port still in
    # TIME_WAIT or a stale process that hasn't released the socket yet.
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        try:
            uvicorn.run(app, host="0.0.0.0", port=port, log_level="info",
                        timeout_keep_alive=30, ws_max_size=16777216,
                        http="httptools")
            break  # Normal shutdown
        except OSError as e:
            if ("10048" in str(e) or "Address already in use" in str(e)):
                if attempt < max_retries:
                    logger.warning(
                        f"Port {port} busy (attempt {attempt}/{max_retries})."
                        f" Retrying in {attempt * 3}s..."
                    )
                    time.sleep(attempt * 3)
                else:
                    logger.error(f"Port {port} still busy after {max_retries} attempts. Exiting.")
                    sys.exit(1)
            else:
                logger.critical(f"FATAL OS ERROR: {e}", exc_info=True)
                raise
        except Exception as e:
            logger.critical(f"FATAL CRASH: {e}", exc_info=True)
            raise
